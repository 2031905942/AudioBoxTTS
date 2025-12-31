import os.path
from pathlib import Path
from typing import Optional

from PySide6.QtCore import Signal

import main
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from Source.Job.base_job import BaseJob
from Source.Utility import wproj_utility
from Source.Utility.config_utility import config_utility, ProjectData
from Source.Utility.voice_excel_utility import VoiceExcelUtility
from waapi import WaapiClient

ACTOR_MIXER_HIERARCHY = "Actor-Mixer Hierarchy"
INTERACTIVE_MUSIC_HIERARCHY = "Interactive Music Hierarchy"
EVENTS = "Events"
SOUNDBANKS = "SoundBanks"
AUTO_IMPORT_WWISE_OBJECT_TYPE_LIST = [ACTOR_MIXER_HIERARCHY, EVENTS, SOUNDBANKS]
CONTAINER_OBJECT_TYPE_DICT = {
    "RD": "RandomSequenceContainer",
    "SQ": "RandomSequenceContainer",
    "SW": "SwitchContainer",
    "BL": "BlendContainer"
}

CONTAINER_FULL_NAME_DICT = {
    "RD": "Random Container",
    "SQ": "Sequence Container",
    "SW": "Switch Container",
    "BL": "Blend Container"
}


class WaapiUtility(VoiceExcelUtility):

    # 信号定义
    validate_wwise_project_succeeded_signal = Signal()
    validate_for_import_succeeded_signal = Signal(str, bool)
    show_wwise_object_check_window_signal = Signal(str, list)
    show_check_window_signal = Signal(str, list)

    def __init__(self, base_job: BaseJob):
        super().__init__(base_job)
        self.client: WaapiClient | None = None

    def connect_waapi(self) -> bool:
        self.update_progress_text_signal.emit("连接Waapi...")
        result: bool = False
        if self.is_connected():
            return True

        try:
            client = WaapiClient()
            if client:
                self.client = client
            result = self.is_connected()
        except Exception as error:
            pass
            # self.error_signal.emit(f"连接Waapi发生异常:\n{error}.")
            # self._print_log_error(f"连接Waapi发生异常:\n{traceback.format_exc()}")
        return result

    def disconnect_waapi(self):
        if self.is_connected():
            self.client.disconnect()
        self.client = None

    def is_connected(self) -> bool:
        result: bool = False
        if self.client:
            result = self.client.is_connected()
        return result

    def is_current_wwise_project_open(self, project_id: str) -> bool:
        self.update_progress_text_signal.emit("检查当前Wwise工程是否开启...")
        project_data: dict[str, str] = config_utility.get_project_data(project_id)
        wwise_project_path: str = project_data.get(ProjectData.WWISE_PROJECT_PATH)
        if not wwise_project_path or not os.path.isfile(wwise_project_path):
            return False
        result = self._get_wwise_object("\"\\\"", ["filePath"])
        if not result:
            return False
        open_wwise_project_path: str = result[0]["filePath"]
        if not main.is_windows_os():
            if open_wwise_project_path.startswith("Z:"):
                open_wwise_project_path = open_wwise_project_path[2:]
            elif open_wwise_project_path.startswith("Y:"):
                open_wwise_project_path = f"{os.path.expanduser('~')}{open_wwise_project_path[2:]}"
        open_wwise_project_path = open_wwise_project_path.replace("\\", "/")
        print(f"当前打开的Wwise工程: \"{open_wwise_project_path}\".")
        if open_wwise_project_path == wwise_project_path:
            return True
        return False

    def is_current_wwise_project_dirty(self) -> bool:
        self.update_progress_text_signal.emit("检查当前Wwise工程是否存在未保存的修改...")
        result = self._get_wwise_object("\"\\\"", ["workunitIsDirty"])
        if not result:
            return False
        for wproj in result:
            wproj_is_dirty: bool = wproj["workunitIsDirty"]
            if wproj_is_dirty:
                return wproj_is_dirty

        result = self._get_wwise_object("select descendants where type = \"WorkUnit\"", ["workunitIsDirty"])
        if not result:
            return False
        for wproj in result:
            wproj_is_dirty: bool = wproj["workunitIsDirty"]
            if wproj_is_dirty:
                return wproj_is_dirty

        return False

    def save_wwise_project(self):
        self.update_progress_text_signal.emit("保存当前Wwise工程...")
        if self.is_connected():
            self.client.call("ak.wwise.core.project.save")

    def reopen_wwise_project(self):
        result = self._get_wwise_object("\"\\\"", ["filePath"])
        if not result:
            return
        open_wwise_project_path: str = result[0]["filePath"]
        args = {
            "path": open_wwise_project_path
        }
        self.client.call("ak.wwise.ui.project.open", args)

    def close_wwise_project(self):
        if self.is_connected():
            self.client.call("ak.wwise.ui.project.close")

    def get_custom_auto_import_sample_type_dict(self) -> dict:
        self.update_progress_text_signal.emit("获取自定义入库素材类型...")
        custom_import_sample_type_list = {}
        physical_folder_parent_path = f"{ACTOR_MIXER_HIERARCHY}"
        result = self._get_wwise_object(f"from object \"{physical_folder_parent_path}\" select children where workunitType = \"folder\"")
        if not result:
            return custom_import_sample_type_list
        for wwise_object in result:
            sample_type_name: str = wwise_object["name"]
            sample_type_abbr: str = wwise_object["notes"]
            if sample_type_name in wproj_utility.IMPORT_CHARACTER_TYPE_DICT.values() or not sample_type_abbr:
                continue
            work_unit_path = f"{ACTOR_MIXER_HIERARCHY}/{sample_type_name}/{sample_type_name}"
            actor_mixer_path = f"{work_unit_path}/{sample_type_name}"
            example_actor_mixer_path = f"{actor_mixer_path}/0 Example"
            event_folder_path = f"{EVENTS}/{sample_type_name}"
            soundbank_folder_path = f"{SOUNDBANKS}/{sample_type_name}"
            result = self._get_wwise_object(f"\"{work_unit_path}\"")
            if result and result[0]["type"] == "WorkUnit":
                result = self._get_wwise_object(f"\"{actor_mixer_path}\"")
                if result and result[0]["type"] == "ActorMixer":
                    result = self._get_wwise_object(f"\"{example_actor_mixer_path}\"")
                    if result and result[0]["type"] == "ActorMixer":
                        custom_import_sample_type_list[sample_type_abbr] = sample_type_name
        return custom_import_sample_type_list

    def validate_for_import_job(self, dir_path: str):
        self.update_progress_text_signal.emit("收集素材...")
        sample_path_list = self.get_files(dir_path, [".wav"])
        exist_sample_type_dict = {}

        # 先收集Wwise工程中创建的符合自动入库规范的
        check_list = []
        has_voice_sample = False
        physical_folder_invalid_check_list = (f"物理文件夹(Physical Folder)对象不存在, 请先在指定路径创建物理文件夹对象", False, [])
        example_actor_mixer_invalid_check_list = (f"模板角色混音(Actor-Mixer)对象不存在, 请先在指定路径创建模板角色混音对象", False, [])
        example_actor_mixer_notes_invalid_check_list = (f"模板角色混音(Actor-Mixer)对象备注不合法, 请为对应的模板角色混音对象备注入库素材时使用的声音类型(\"SFX\"或\"Voice\")", False, [])
        for sample_path in sample_path_list:
            file_name_stem = Path(sample_path).stem
            file_name_split = file_name_stem.split()
            sample_type = file_name_split[0]
            if sample_type in wproj_utility.IMPORT_STORY_VOICE_TYPE_LIST:
                has_voice_sample = True
            if sample_type in wproj_utility.IMPORT_CHARACTER_TYPE_LIST or sample_type in wproj_utility.CUSTOM_IMPORT_CHARACTER_TYPE_DICT:
                character_sample_type = file_name_split[2]
            else:
                character_sample_type = "Dialog"
            if sample_type not in exist_sample_type_dict.keys():
                exist_sample_type_dict[sample_type] = set()
            exist_sample_type_dict[sample_type].add(character_sample_type)

        for sample_type in exist_sample_type_dict.keys():
            character_sample_type_set = exist_sample_type_dict[sample_type]
            if sample_type in wproj_utility.IMPORT_CHARACTER_TYPE_DICT:
                sample_type = wproj_utility.IMPORT_CHARACTER_TYPE_DICT[sample_type]
            elif sample_type in wproj_utility.CUSTOM_IMPORT_CHARACTER_TYPE_DICT:
                sample_type = wproj_utility.CUSTOM_IMPORT_CHARACTER_TYPE_DICT[sample_type]
            object_color = 0
            for wwise_object_type in AUTO_IMPORT_WWISE_OBJECT_TYPE_LIST:
                physical_folder_path = f"{wwise_object_type}/{sample_type}"
                result = self._get_wwise_object(f"\"{physical_folder_path}\"", ["workunitType"])
                if object_color == 0 and result:
                    current_object_color = result[0].get("@Color", 0)
                    if current_object_color != 0:
                        object_color = current_object_color
                if result and result[0]["workunitType"] == "folder":
                    arg_dict = {
                        "objects": [
                            {
                                "object":         result[0]["id"],
                                "@OverrideColor": object_color,
                                "@Inclusion":     True,
                            }
                        ]
                    }
                    self._set_wwise_object(arg_dict)
                if not result or result[0]["workunitType"] != "folder":
                    physical_folder_invalid_check_str = (f"\"{physical_folder_path}\"", False, None)
                    arg_dict = {
                        "objects": [
                            {
                                "object":   f"\\{wwise_object_type}",
                                "children": [
                                    {
                                        "type":   "Physical Folder",
                                        "name":   sample_type,
                                        "@Color": object_color,
                                    }
                                ]
                            }
                        ]
                    }
                    set_result = self._set_wwise_object(arg_dict)
                    if not set_result:
                        physical_folder_invalid_check_list[2].append(physical_folder_invalid_check_str)

            for character_sample_type in character_sample_type_set:
                example_actor_mixer_path = f"{ACTOR_MIXER_HIERARCHY}/{sample_type}/{sample_type}/{sample_type}/0 Example/{character_sample_type}"
                example_actor_mixer_invalid_check_str = (f"\"{example_actor_mixer_path}\"", False, None)
                result = self._get_wwise_object(f"\"{example_actor_mixer_path}\"")
                if not result:
                    example_actor_mixer_invalid_check_list[2].append(example_actor_mixer_invalid_check_str)
                else:
                    notes = result[0]["notes"]
                    if notes == "Voice":
                        has_voice_sample = True
                    elif not has_voice_sample and notes != "SFX" and notes != "Voice":
                        example_actor_mixer_notes_invalid_check_list[2].append(example_actor_mixer_invalid_check_str)

        if len(physical_folder_invalid_check_list[2]) > 0:
            check_list.append(physical_folder_invalid_check_list)

        if len(example_actor_mixer_invalid_check_list[2]) > 0:
            check_list.append(example_actor_mixer_invalid_check_list)

        if len(example_actor_mixer_notes_invalid_check_list[2]) > 0:
            check_list.append(example_actor_mixer_notes_invalid_check_list)

        if len(check_list) > 0:
            self.show_check_window_signal.emit("Wwise入库条件检查列表", check_list)
            self.finish_signal.emit("结果", "任务中止, 请先排查完Wwise入库条件的问题再执行", "error")
        else:
            self.validate_for_import_succeeded_signal.emit(dir_path, has_voice_sample)
            self.show_result_info_bar_signal.emit("success", "结果", "验证Wwise入库条件成功")

    def import_sample_job(self, parameter: dict):
        dir_path: str = parameter["dir_path"]
        self.backup_directory(dir_path)
        import_language_list: list[str] = parameter.get("import_language_list", [])
        voice_excel_file_path: str = parameter.get("voice_excel_file_path", "")

        check_list = []

        container_type_invalid_check_list = ("现有容器Wwise对象类型与入库素材目标容器类型不一样, 跳过以下素材入库操作", None, [])
        sound_type_invalid_check_list = ("现有声音Wwise对象类型与入库素材目标声音类型不一样, 跳过以下素材入库操作", None, [])
        voice_type_invalid_check_list = ("现有声音Wwise对象的语音类型与入库素材目标语音类型不一样, 跳过以下素材入库操作", None, [])

        sample_skip_imported_dir_path = f"{dir_path}/跳过入库的素材"
        sample_cannot_imported_dir_path = f"{dir_path}/无法入库的素材"
        sample_imported_dir_path = f"{dir_path}/成功入库的素材"

        voice_excel_workbook: Optional[Workbook] = None
        is_voice_excel_workbook_modify = False
        if voice_excel_file_path and len(import_language_list) > 0:
            if not self.check_workbook_writability(voice_excel_file_path):
                return
            voice_excel_workbook = self.load_workbook(voice_excel_file_path)
            if not voice_excel_workbook:
                self.finish_signal.emit("结果", "任务中止", "error")
                return
            if not self.check_worksheet_header_validity(voice_excel_file_path, voice_excel_workbook):
                self.finish_signal.emit("结果", "任务中止", "error")
                return

        voice_excel_event_info_dict = self._read_voice_excel_data_for_import_job(voice_excel_workbook)

        self.update_progress_text_signal.emit("收集素材...")
        sample_path_list = self.get_files(dir_path, [".wav"])
        total_count = len(sample_path_list)
        self.update_progress_total_count_signal.emit(total_count)
        current_count = 0
        import_count = 0
        skip_count = 0
        for sample_path in sample_path_list:
            if self.cancel_job:
                break
            file_name = os.path.basename(sample_path)
            file_name_stem = Path(sample_path).stem
            original_sample_path = sample_path
            if not main.is_windows_os():
                user_path_prefix = os.path.expanduser('~')
                if sample_path.startswith(user_path_prefix):
                    sample_path = f"Y:{sample_path[len(user_path_prefix):]}"
                else:
                    sample_path = f"Z:{sample_path}"
            sample_path = sample_path.replace("/", "\\")
            file_name_split = file_name_stem.split()
            self.update_progress_text_signal.emit(f"素材入库...\n\"{file_name}\"")
            sample_type_abbr = file_name_split[0]
            sample_type = ""
            if sample_type_abbr in wproj_utility.IMPORT_CHARACTER_TYPE_DICT:
                sample_type = wproj_utility.IMPORT_CHARACTER_TYPE_DICT[sample_type_abbr]
            elif sample_type_abbr in wproj_utility.CUSTOM_IMPORT_CHARACTER_TYPE_DICT:
                sample_type = wproj_utility.CUSTOM_IMPORT_CHARACTER_TYPE_DICT[sample_type_abbr]
            character_id = file_name_split[1]

            example_actor_mixer_path = f"{ACTOR_MIXER_HIERARCHY}/{sample_type}/{sample_type}/{sample_type}/0 Example"
            result = self._get_wwise_object(f"\"{example_actor_mixer_path}\"")
            example_actor_mixer_id = result[0]["id"]

            root_actor_mixer_path = f"{ACTOR_MIXER_HIERARCHY}/{sample_type}/{sample_type}/{sample_type}"
            result = self._get_wwise_object(f"\"{root_actor_mixer_path}\"")
            root_actor_mixer_id = result[0]["id"]

            is_container = False
            if "-" in file_name_split[-1]:
                is_container = True

            is_character_sample = False
            is_story_voice_sample = False
            if sample_type_abbr in wproj_utility.IMPORT_CHARACTER_TYPE_LIST or sample_type_abbr in wproj_utility.CUSTOM_IMPORT_CHARACTER_TYPE_DICT:
                is_character_sample = True
            elif sample_type_abbr in wproj_utility.IMPORT_STORY_VOICE_TYPE_LIST:
                is_story_voice_sample = True

            if is_character_sample:
                character_sample_type = file_name_split[2]
            elif is_story_voice_sample:
                character_sample_type = "Dialog"
            else:
                character_sample_type = file_name_split[2]
            # 从模板对象中的备注确认素材以"SFX"还是"Voice"的形式入库
            example_character_sample_type_actor_mixer_path = f"{example_actor_mixer_path}/{character_sample_type}"
            result = self._get_wwise_object(f"\"{example_character_sample_type_actor_mixer_path}\"")
            import_type: str = result[0]["notes"]
            import_as_voice = import_type == "Voice"
            if import_as_voice and len(import_language_list) == 0:
                dst_sample_path: str = f"{sample_skip_imported_dir_path}/{file_name}"
                self.move_file(original_sample_path, dst_sample_path)
                skip_count += 1
                current_count += 1
                self.update_progress_current_count_signal.emit(current_count)
                continue

            if sample_type_abbr in wproj_utility.IMPORT_CHARACTER_TYPE_LIST or sample_type_abbr in wproj_utility.CUSTOM_IMPORT_CHARACTER_TYPE_DICT:
                # 素材是角色素材的情况
                character_id_work_unit_name = f"{sample_type} {character_id}"
                character_id_work_unit_path = f"{root_actor_mixer_path}/{character_id_work_unit_name}"

                result = self._get_wwise_object(f"\"{character_id_work_unit_path}\"")
                if not result:
                    result = self._create_wwise_object(root_actor_mixer_id, "WorkUnit", character_id_work_unit_name)
                    character_id_work_unit_id: str = result["id"]
                else:
                    character_id_work_unit_id: str = result[0]["id"]

                # 如果角色旧入库规范的路径的角色混音对象存在, 则移动到新规范的工作单元中
                old_character_id_actor_mixer_path = f"{root_actor_mixer_path}/{character_id}"
                result = self._get_wwise_object(f"\"{old_character_id_actor_mixer_path}\"")
                if result and result[0]["type"] == "ActorMixer":
                    self._move_wwise_object(result[0]["id"], character_id_work_unit_id)

                character_id_actor_mixer_path = f"{character_id_work_unit_path}/{character_id}"
                # 如果新规范下的角色的角色混音对象不存在, 则从模板对象复制过来
                result = self._get_wwise_object(f"\"{character_id_actor_mixer_path}\"")
                if not result:
                    result = self._copy_wwise_object(example_actor_mixer_id, character_id_work_unit_id)
                    self._rename_wwise_object(result["id"], character_id)

                # 确认入库的声音对象所在的角色混音对象
                example_child_actor_mixer_path = example_actor_mixer_path
                import_parent_object_path = f"{character_id_actor_mixer_path}"
                result = self._get_wwise_object(f"\"{import_parent_object_path}\"")
                import_parent_object_id = result[0]["id"]
                for i in range(2, len(file_name_split)):
                    child_actor_mixer_name = file_name_split[i]
                    check_import_actor_mixer_path = f"{import_parent_object_path}/{child_actor_mixer_name}"
                    check_example_child_actor_mixer_path = f"{example_child_actor_mixer_path}/{child_actor_mixer_name}"
                    result = self._get_wwise_object(f"\"{check_import_actor_mixer_path}\"")
                    if result:
                        import_parent_object_path = check_import_actor_mixer_path
                        example_child_actor_mixer_path = check_example_child_actor_mixer_path
                        import_parent_object_id = result[0]["id"]
                    else:
                        result = self._get_wwise_object(f"\"{check_example_child_actor_mixer_path}\"")
                        if result:
                            example_child_actor_mixer_id = result[0]["id"]
                            result = self._copy_wwise_object(example_child_actor_mixer_id, import_parent_object_id)
                            if result:
                                import_parent_object_path = check_import_actor_mixer_path
                                import_parent_object_id = result["id"]
                        else:
                            break

                event_sound_object_id = ""
                switch_group_path = ""
                switch_name = ""
                if is_container:
                    # 如果入库素材是容器命名格式, 则进行容器处理
                    switch_group_id = ""
                    container_element_split = file_name_split[-1].split("-")
                    container_type_abbr = container_element_split[0]
                    if container_type_abbr == "SW":
                        switch_name = container_element_split[1]
                    container_type = CONTAINER_OBJECT_TYPE_DICT[container_type_abbr]
                    container_name = " ".join(file_name_split[:-1])
                    container_path = f"{import_parent_object_path}/{container_name}"
                    import_parent_object_path = container_path
                    result = self._get_wwise_object(f"\"{container_path}\"", ["@RandomOrSequence", "@SwitchGroupOrStateGroup"])
                    if result:
                        # 如果容器已经存在, 则进行容器类型检查
                        exist_container_type = result[0]["type"]
                        import_parent_object_id = result[0]["id"]
                        event_sound_object_id = import_parent_object_id
                        if exist_container_type == "RandomSequenceContainer":
                            if result[0]["@RandomOrSequence"] == 0:
                                exist_container_type_abbr = "SQ"
                            else:
                                exist_container_type_abbr = "RD"
                            if exist_container_type_abbr != container_type_abbr:
                                # 如果现有容器类型与入库素材目标容器类型不一样, 则跳过该素材
                                container_type_invalid_check_str = (
                                    f"素材: \"{file_name}\"; 现有容器的Wwise对象类型: \"{CONTAINER_FULL_NAME_DICT[exist_container_type_abbr]}\"",
                                    import_parent_object_id, None)
                                container_type_invalid_check_list[2].append(container_type_invalid_check_str)
                                dst_sample_path: str = f"{sample_cannot_imported_dir_path}/{file_name}"
                                self.move_file(original_sample_path, dst_sample_path)
                                skip_count += 1
                                current_count += 1
                                self.update_progress_current_count_signal.emit(current_count)
                                continue
                        elif container_type != exist_container_type:
                            # 如果现有对象类型与入库素材目标容器类型不一样, 则跳过该素材
                            container_type_invalid_check_str = (
                                f"素材: \"{file_name}\"; 现有容器的Wwise对象类型: \"{exist_container_type}\"",
                                import_parent_object_id, None)
                            container_type_invalid_check_list[2].append(container_type_invalid_check_str)
                            dst_sample_path: str = f"{sample_cannot_imported_dir_path}/{file_name}"
                            self.move_file(original_sample_path, dst_sample_path)
                            skip_count += 1
                            current_count += 1
                            self.update_progress_current_count_signal.emit(current_count)
                            continue
                        elif exist_container_type == "SwitchContainer":
                            switch_group_id = result[0]["@SwitchGroupOrStateGroup"]["id"]
                    else:
                        # 如果容器不存在, 则确认模板容器是否存在
                        example_container_path = f"{example_child_actor_mixer_path}/0 Example {CONTAINER_FULL_NAME_DICT[container_type_abbr]}"
                        result = self._get_wwise_object(f"\"{example_container_path}\"")
                        if result:
                            # 如果模板容器存在, 则复制并重命名到入库路径中
                            result = self._copy_wwise_object(result[0]["id"], import_parent_object_id, "rename")
                            import_parent_object_id = result["id"]
                            event_sound_object_id = import_parent_object_id
                            self._rename_wwise_object(event_sound_object_id, container_name)
                            if container_type_abbr == "SW":
                                result = self._get_wwise_object(f"\"{event_sound_object_id}\"", ["@SwitchGroupOrStateGroup"])
                                switch_group_id = result[0]["@SwitchGroupOrStateGroup"]
                        else:
                            # 如果模板容器不存在, 则新建容器
                            arg_dict: Optional[dict] = None
                            if container_type_abbr == "RD":
                                arg_dict = {
                                    "@RandomOrSequence": 1
                                }
                            elif container_type_abbr == "SQ":
                                arg_dict = {
                                    "@RandomOrSequence": 0
                                }
                            result = self._create_wwise_object(import_parent_object_id, container_type, container_name, arg_dict)
                            import_parent_object_id = result["id"]
                            event_sound_object_id = import_parent_object_id

                    if switch_group_id:
                        result = self._get_wwise_object(f"\"{switch_group_id}\"")
                        switch_group_path = result[0]["path"]

                import_sound_path = f"{import_parent_object_path}/{file_name_stem}"
                result = self._get_wwise_object(f"\"{import_sound_path}\"", ["@IsVoice"])
                if result:
                    # 如果入库声音对象已经存在, 则进行类型检查
                    import_sound_id = result[0]["id"]
                    import_sound_object_type = result[0]["type"]
                    is_voice = result[0].get("@IsVoice", False)
                    if import_sound_object_type != "Sound":
                        # 如果现有入库声音对象的类型不是"Sound"(即非声音类型), 则跳过该素材
                        sound_type_invalid_check_str = (
                            f"素材: \"{file_name}\"; 现有声音的Wwise对象类型: \"{import_sound_object_type}\"", result[0]["id"], None)
                        sound_type_invalid_check_list[2].append(sound_type_invalid_check_str)
                        dst_sample_path: str = f"{sample_cannot_imported_dir_path}/{file_name}"
                        self.move_file(original_sample_path, dst_sample_path)
                        skip_count += 1
                        current_count += 1
                        self.update_progress_current_count_signal.emit(current_count)
                        continue
                    elif import_as_voice != is_voice:
                        # 如果现有入库声音对象的语言类型不同("SFX"和"Voice"), 则跳过该素材
                        voice_type_invalid_check_str = (
                            f"素材: \"{file_name}\"; 需要入库为语音: \"{import_as_voice}\"; 现有声音的Wwise对象是否语音: \"{is_voice}\"",
                            result[0]["id"], None)
                        voice_type_invalid_check_list[2].append(voice_type_invalid_check_str)
                        dst_sample_path: str = f"{sample_cannot_imported_dir_path}/{file_name}"
                        self.move_file(original_sample_path, dst_sample_path)
                        skip_count += 1
                        current_count += 1
                        self.update_progress_current_count_signal.emit(current_count)
                        continue
                else:
                    result = self._create_wwise_object(import_parent_object_id, "Sound", file_name_stem, {
                        "@IsVoice": import_as_voice
                    })
                    import_sound_id = result["id"]

                if not is_container:
                    event_sound_object_id = import_sound_id

                # 入库操作
                import_audio_path = import_parent_object_path
                import_audio_path_split = import_audio_path.split("/")
                import_audio_path = "/".join(import_audio_path_split[5:])
                import_audio_path = f"{import_audio_path_split[3]}/{import_audio_path}"
                if import_as_voice:
                    for import_language in import_language_list:
                        self._import_audio(f"<AudioFileSource>{file_name_stem} ({import_language})", sample_path, import_language, import_sound_id, import_audio_path)
                else:
                    self._import_audio(f"<AudioFileSource>{file_name_stem}", sample_path, root=import_sound_id, import_path=import_audio_path)

                if switch_group_path:
                    switch_path = f"{switch_group_path}\\{switch_name}"
                    result = self._get_wwise_object(f"\"{switch_path}\"")
                    if result:
                        switch_id = result[0]["id"]
                        self._add_switch_container_assignment(import_sound_id, switch_id)

                # 事件层级操作
                event_folder_path = f"{EVENTS}/{sample_type}"
                result = self._get_wwise_object(f"\"{event_folder_path}\"")
                event_root_folder_id = result[0]["id"]
                result = self._create_wwise_object(event_root_folder_id, "WorkUnit", character_id, {
                    "onNameConflict": "merge"
                })
                character_event_work_unit_id = result["id"]

                # 旧入库规范的文件夹迁移至新规范
                old_character_event_folder_path = f"{event_folder_path}/{character_id}/{character_id}"
                result = self._get_wwise_object(f"\"{old_character_event_folder_path}\"")
                if result and result[0]["type"] == "Folder":
                    old_character_event_folder_id = result[0]["id"]
                    result = self._get_wwise_object(f"\"{old_character_event_folder_path}\" select children")
                    for child_event_folder in result:
                        self._move_wwise_object(child_event_folder["id"], character_event_work_unit_id)
                    result = self._get_wwise_object(f"\"{SOUNDBANKS}/{sample_type}\" select descendants where name = \"{sample_type}{character_id}\" and type = \"SoundBank\"")
                    if result:
                        character_soundbank_id = result[0]["id"]
                        soundbank_inclusions_dict = {
                            "object": old_character_event_folder_id,
                            "filter": ["events", "structures", "media"]
                        }
                        self._set_soundbank_inclusions(character_soundbank_id, [soundbank_inclusions_dict], "remove")
                    self._delete_wwise_object(old_character_event_folder_id)

                character_event_folder_split = import_audio_path_split[6:]
                if is_container:
                    character_event_folder_split = character_event_folder_split[:-1]
                event_folder_parent_id = character_event_work_unit_id
                for character_event_folder_name in character_event_folder_split:
                    result = self._create_wwise_object(event_folder_parent_id, "Folder", character_event_folder_name, {
                        "onNameConflict": "merge"
                    })
                    event_folder_parent_id = result["id"]
                if is_container:
                    event_name = "_".join(file_name_split[:-1])
                else:
                    event_name = file_name_stem.replace(" ", "_")
                result = self._get_wwise_object(f"where name = \"{event_name}\" and type = \"Event\"")
                if result:
                    event_id = result[0]["id"]
                    if result[0]["parent"]["id"] != event_folder_parent_id:
                        self._move_wwise_object(event_id, event_folder_parent_id)
                    result = self._get_wwise_object(f"\"{event_id}\" select children")
                    if not result:
                        action_arg_dict = {
                            "@ActionType": 1,
                            "@Target":     event_sound_object_id
                        }
                        self._create_wwise_object(event_id, "Action", "", action_arg_dict)
                    else:
                        action_id = result[0]['id']
                        self._set_wwise_object_property(action_id, "ActionType", 1)
                        self._set_wwise_object_reference(action_id, "Target", event_sound_object_id)
                else:
                    event_children_arg_dict = {
                        "children": [
                            {
                                "type":        "Action",
                                "name":        "",
                                "@ActionType": 1,
                                "@Target":     event_sound_object_id
                            }
                        ]
                    }
                    self._create_wwise_object(event_folder_parent_id, "Event", event_name, event_children_arg_dict)
                if import_as_voice:
                    if self._clear_voice_excel_dialogue_record_state_for_import_job(voice_excel_workbook, voice_excel_event_info_dict, event_name):
                        is_voice_excel_workbook_modify = True

                # 声音库层级操作
                soundbank_folder_path = f"{SOUNDBANKS}/{sample_type}"
                result = self._get_wwise_object(f"\"{soundbank_folder_path}\"")
                soundbank_folder_id = result[0]["id"]
                result = self._create_wwise_object(soundbank_folder_id, "WorkUnit", character_id, {
                    "onNameConflict": "merge"
                })
                character_soundbank_work_unit_id = result["id"]
                soundbank_name = f"{sample_type}{character_id}"
                result = self._get_wwise_object(f"where name = \"{soundbank_name}\" and type = \"SoundBank\"")
                if result:
                    soundbank_id = result[0]["id"]
                    soundbank_parent_id = result[0]["parent"]["id"]
                    if soundbank_parent_id != character_soundbank_work_unit_id:
                        self._move_wwise_object(soundbank_id, character_soundbank_work_unit_id)
                else:
                    result = self._create_wwise_object(character_soundbank_work_unit_id, "SoundBank", soundbank_name)
                    soundbank_id = result["id"]
                soundbank_inclusions_dict = {
                    "object": character_event_work_unit_id,
                    "filter": ["events", "structures", "media"]
                }
                self._set_soundbank_inclusions(soundbank_id, [soundbank_inclusions_dict], "replace")

            elif sample_type_abbr in wproj_utility.IMPORT_STORY_VOICE_TYPE_LIST:
                story_work_unit_name = sample_type
                current_parent_path = root_actor_mixer_path
                current_parent_id = root_actor_mixer_id
                import_audio_path = sample_type
                container_element_index = 0
                if is_container:
                    container_element_index = -1

                # 也支持纯数字ID形式的语音素材, 比如"VO 100100123", 默认的策略是左数从头到倒数第三位数字会划分为工作单元和角色混音容器, 事件层级的工作单元和文件夹, 以及声音库, 比如"VO 1001001XX"的语音素材都会归到"STO 1001001"的工作单元及角色混音容器, 声音库"STO_1001001"
                is_full_id_voice = False
                full_id_work_unit_name: Optional[str] = None
                if len(file_name_split) + container_element_index == 2 and file_name_split[1].isdigit():
                    is_full_id_voice = True

                if is_full_id_voice:
                    full_id_work_unit_name = str(int(file_name_split[1]) // 100)

                for i in range(1, len(file_name_split) - 1 + container_element_index):
                    file_name_element = file_name_split[i]

                    story_work_unit_name = f"{story_work_unit_name} {file_name_element}"
                    result = self._create_wwise_object(current_parent_id, "WorkUnit", story_work_unit_name, {
                        "onNameConflict": "merge"
                    })
                    current_parent_id = result["id"]
                    import_audio_path = f"{import_audio_path}/{file_name_element}"

                    current_parent_path = f"{current_parent_path}/{story_work_unit_name}"
                    result = self._get_wwise_object(f"\"{current_parent_path}/{file_name_element}\"")
                    if result:
                        current_parent_id = result[0]["id"]
                    else:
                        result = self._create_wwise_object(current_parent_id, "ActorMixer", file_name_element)
                        current_parent_id = result["id"]
                    current_parent_path = f"{current_parent_path}/{file_name_element}"

                if is_full_id_voice:
                    story_work_unit_name = f"{story_work_unit_name} {full_id_work_unit_name}"
                    result = self._create_wwise_object(current_parent_id, "WorkUnit", story_work_unit_name, {
                        "onNameConflict": "merge"
                    })
                    current_parent_id = result["id"]
                    import_audio_path = f"{import_audio_path}/{full_id_work_unit_name}"

                    current_parent_path = f"{current_parent_path}/{story_work_unit_name}"
                    result = self._get_wwise_object(f"\"{current_parent_path}/{full_id_work_unit_name}\"")
                    if result:
                        current_parent_id = result[0]["id"]
                    else:
                        result = self._create_wwise_object(current_parent_id, "ActorMixer", full_id_work_unit_name)
                        current_parent_id = result["id"]
                    current_parent_path = f"{current_parent_path}/{full_id_work_unit_name}"

                import_audio_path = f"{import_audio_path}/Dialog"
                sfx_actor_mixer_path = f"{current_parent_path}/SFX"
                result = self._get_wwise_object(f"\"{sfx_actor_mixer_path}\"")
                if not result:
                    result = self._get_wwise_object(f"\"{example_actor_mixer_path}/SFX\"")
                    if result:
                        self._copy_wwise_object(result[0]["id"], current_parent_id)
                    else:
                        self._create_wwise_object(current_parent_id, "ActorMixer", "SFX")

                dialog_actor_mixer_path = f"{current_parent_path}/Dialog"
                result = self._get_wwise_object(f"\"{dialog_actor_mixer_path}\"")
                if not result:
                    result = self._get_wwise_object(f"\"{example_actor_mixer_path}/Dialog\"")
                    if result:
                        result = self._copy_wwise_object(result[0]["id"], current_parent_id)
                    else:
                        result = self._create_wwise_object(current_parent_id, "ActorMixer", "Dialog")
                    current_parent_id = result["id"]
                else:
                    current_parent_id = result[0]["id"]
                current_parent_path = dialog_actor_mixer_path

                import_parent_object_path = current_parent_path
                import_parent_object_id = current_parent_id

                event_sound_object_id = ""
                switch_group_path = ""
                switch_name = ""

                if is_container:
                    # 如果入库素材是容器命名格式, 则进行容器处理
                    switch_group_id = ""
                    container_element_split = file_name_split[-1].split("-")
                    container_type_abbr = container_element_split[0]
                    if container_type_abbr == "SW":
                        switch_name = container_element_split[1]
                    container_type = CONTAINER_OBJECT_TYPE_DICT[container_type_abbr]
                    container_name = " ".join(file_name_split[:-1])
                    container_path = f"{import_parent_object_path}/{container_name}"
                    import_audio_path = f"{import_audio_path}/{container_name}"
                    import_parent_object_path = container_path
                    result = self._get_wwise_object(f"\"{container_path}\"", ["@RandomOrSequence", "@SwitchGroupOrStateGroup"])
                    if result:
                        # 如果容器已经存在, 则进行容器类型检查
                        exist_container_type = result[0]["type"]
                        import_parent_object_id = result[0]["id"]
                        event_sound_object_id = import_parent_object_id
                        if exist_container_type == "RandomSequenceContainer":
                            if result[0]["@RandomOrSequence"] == 0:
                                exist_container_type_abbr = "SQ"
                            else:
                                exist_container_type_abbr = "RD"
                            if exist_container_type_abbr != container_type_abbr:
                                # 如果现有容器类型与入库素材目标容器类型不一样, 则跳过该素材
                                container_type_invalid_check_str = (
                                    f"素材: \"{file_name}\"; 现有容器的Wwise对象类型: \"{CONTAINER_FULL_NAME_DICT[exist_container_type_abbr]}\"",
                                    import_parent_object_id, None)
                                container_type_invalid_check_list[2].append(container_type_invalid_check_str)
                                dst_sample_path: str = f"{sample_cannot_imported_dir_path}/{file_name}"
                                self.move_file(original_sample_path, dst_sample_path)
                                skip_count += 1
                                current_count += 1
                                self.update_progress_current_count_signal.emit(current_count)
                                continue
                        elif container_type != exist_container_type:
                            # 如果现有对象类型与入库素材目标容器类型不一样, 则跳过该素材
                            container_type_invalid_check_str = (
                                f"素材: \"{file_name}\"; 现有容器的Wwise对象类型: \"{exist_container_type}\"",
                                import_parent_object_id, None)
                            container_type_invalid_check_list[2].append(container_type_invalid_check_str)
                            dst_sample_path: str = f"{sample_cannot_imported_dir_path}/{file_name}"
                            self.move_file(original_sample_path, dst_sample_path)
                            skip_count += 1
                            current_count += 1
                            self.update_progress_current_count_signal.emit(current_count)
                            continue
                        elif exist_container_type == "SwitchContainer":
                            switch_group_id = result[0]["@SwitchGroupOrStateGroup"]["id"]
                    else:
                        # 如果容器不存在, 则确认模板容器是否存在
                        example_container_path = f"{example_actor_mixer_path}/Dialog/0 Example {CONTAINER_FULL_NAME_DICT[container_type_abbr]}"
                        result = self._get_wwise_object(f"\"{example_container_path}\"")
                        if result:
                            # 如果模板容器存在, 则复制并重命名到入库路径中
                            result = self._copy_wwise_object(result[0]["id"], import_parent_object_id, "rename")
                            import_parent_object_id = result["id"]
                            event_sound_object_id = import_parent_object_id
                            self._rename_wwise_object(result["id"], container_name)
                            if container_type_abbr == "SW":
                                result = self._get_wwise_object(f"\"{result['id']}\"", ["@SwitchGroupOrStateGroup"])
                                switch_group_id = result[0]["@SwitchGroupOrStateGroup"]
                        else:
                            # 如果模板容器不存在, 则新建容器
                            arg_dict: Optional[dict] = None
                            if container_type_abbr == "RD":
                                arg_dict = {
                                    "@RandomOrSequence": 1
                                }
                            elif container_type_abbr == "SQ":
                                arg_dict = {
                                    "@RandomOrSequence": 0
                                }
                            result = self._create_wwise_object(import_parent_object_id, container_type, container_name, arg_dict)
                            import_parent_object_id = result["id"]
                            event_sound_object_id = import_parent_object_id

                    if switch_group_id:
                        result = self._get_wwise_object(f"\"{switch_group_id}\"")
                        switch_group_path = result[0]["path"]

                import_sound_path = f"{import_parent_object_path}/{file_name_stem}"
                result = self._get_wwise_object(f"\"{import_sound_path}\"", ["@IsVoice"])
                if result:
                    # 如果入库声音对象已经存在, 则进行类型检查
                    import_sound_id = result[0]["id"]
                    import_sound_object_type = result[0]["type"]
                    is_voice = result[0].get("@IsVoice", False)
                    if import_sound_object_type != "Sound":
                        # 如果现有入库声音对象的类型不是"Sound"(即非声音类型), 则跳过该素材
                        sound_type_invalid_check_str = (
                            f"素材: \"{file_name}\"; 现有声音的Wwise对象类型: \"{import_sound_object_type}\"", result[0]["id"], None)
                        sound_type_invalid_check_list[2].append(sound_type_invalid_check_str)
                        dst_sample_path: str = f"{sample_cannot_imported_dir_path}/{file_name}"
                        self.move_file(original_sample_path, dst_sample_path)
                        skip_count += 1
                        current_count += 1
                        self.update_progress_current_count_signal.emit(current_count)
                        continue
                    elif not is_voice:
                        # 如果现有入库声音对象的语言类型不同("SFX"和"Voice"), 则跳过该素材
                        voice_type_invalid_check_str = (
                            f"素材: \"{file_name}\"; 需要入库为语音: \"True\"; 现有声音的Wwise对象是否语音: \"False\"",
                            result[0]["id"], None)
                        voice_type_invalid_check_list[2].append(voice_type_invalid_check_str)
                        dst_sample_path: str = f"{sample_cannot_imported_dir_path}/{file_name}"
                        self.move_file(original_sample_path, dst_sample_path)
                        skip_count += 1
                        current_count += 1
                        self.update_progress_current_count_signal.emit(current_count)
                        continue
                else:
                    result = self._create_wwise_object(import_parent_object_id, "Sound", file_name_stem, {
                        "@IsVoice": True
                    })
                    import_sound_id = result["id"]

                if not is_container:
                    event_sound_object_id = import_sound_id

                # 入库操作
                for import_language in import_language_list:
                    self._import_audio(f"<AudioFileSource>{file_name_stem} ({import_language})", sample_path, import_language, import_sound_id, import_audio_path)

                if switch_group_path:
                    switch_path = f"{switch_group_path}\\{switch_name}"
                    result = self._get_wwise_object(f"\"{switch_path}\"")
                    if result:
                        switch_id = result[0]["id"]
                        self._add_switch_container_assignment(import_sound_id, switch_id)

                # 事件层级操作
                event_folder_path = f"{EVENTS}/{sample_type}"
                result = self._get_wwise_object(f"\"{event_folder_path}\"")
                current_parent_id = result[0]["id"]
                story_event_work_unit_name = ""

                for i in range(1, len(file_name_split) - 1 + container_element_index):
                    file_name_element = file_name_split[i]

                    if not story_event_work_unit_name:
                        story_event_work_unit_name = f"{file_name_element}"
                    else:
                        story_event_work_unit_name = f"{story_event_work_unit_name} {file_name_element}"
                    result = self._get_wwise_object(f"from object \"{EVENTS}/{sample_type}\" select descendants where name = \"{story_event_work_unit_name}\" and type = \"WorkUnit\"")
                    if result and result[0]["parent"]["id"] != current_parent_id:
                        self._move_wwise_object(result[0]["id"], current_parent_id)
                        current_parent_id = result[0]["id"]
                    else:
                        result = self._create_wwise_object(current_parent_id, "WorkUnit", story_event_work_unit_name, {
                            "onNameConflict": "merge"
                        })
                        current_parent_id = result["id"]

                if is_full_id_voice:
                    story_event_work_unit_name = full_id_work_unit_name
                    result = self._create_wwise_object(current_parent_id, "WorkUnit", story_event_work_unit_name, {
                        "onNameConflict": "merge"
                    })
                    current_parent_id = result["id"]

                result = self._create_wwise_object(current_parent_id, "Folder", "SFX", {
                    "onNameConflict": "merge"
                })
                sfx_event_folder_id = result["id"]
                result = self._create_wwise_object(current_parent_id, "Folder", "Dialog", {
                    "onNameConflict": "merge"
                })
                dialog_event_folder_id = result["id"]
                current_parent_id = dialog_event_folder_id

                if is_container:
                    event_name = "_".join(file_name_split[:-1])
                else:
                    event_name = file_name_stem.replace(" ", "_")
                result = self._get_wwise_object(f"where name = \"{event_name}\" and type = \"Event\"")
                if result:
                    event_id = result[0]["id"]
                    if result[0]["parent"]["id"] != current_parent_id:
                        self._move_wwise_object(event_id, current_parent_id)
                    result = self._get_wwise_object(f"\"{event_id}\" select children")
                    if not result:
                        action_arg_dict = {
                            "@ActionType": 1,
                            "@Target":     event_sound_object_id
                        }
                        self._create_wwise_object(event_id, "Action", "", action_arg_dict)
                    else:
                        action_id = result[0]['id']
                        self._set_wwise_object_property(action_id, "ActionType", 1)
                        self._set_wwise_object_reference(action_id, "Target", event_sound_object_id)
                else:
                    event_children_arg_dict = {
                        "children": [
                            {
                                "type":        "Action",
                                "name":        "",
                                "@ActionType": 1,
                                "@Target":     event_sound_object_id
                            }
                        ]
                    }
                    self._create_wwise_object(current_parent_id, "Event", event_name, event_children_arg_dict)
                if self._clear_voice_excel_dialogue_record_state_for_import_job(voice_excel_workbook, voice_excel_event_info_dict, event_name):
                    is_voice_excel_workbook_modify = True

                # 声音库层级操作
                soundbank_folder_path = f"{SOUNDBANKS}/{sample_type}"
                result = self._get_wwise_object(f"\"{soundbank_folder_path}\"")
                current_parent_id = result[0]["id"]
                current_parent_path = soundbank_folder_path
                story_soundbank_work_unit_name = ""

                for i in range(1, len(file_name_split) - 1 + container_element_index):
                    file_name_element = file_name_split[i]

                    if not story_soundbank_work_unit_name:
                        story_soundbank_work_unit_name = f"{file_name_element}"
                    else:
                        story_soundbank_work_unit_name = f"{story_soundbank_work_unit_name} {file_name_element}"
                    result = self._get_wwise_object(f"from object \"{SOUNDBANKS}/{sample_type}\" select descendants where name = \"{story_soundbank_work_unit_name}\" and type = \"WorkUnit\"")
                    if result and result[0]["parent"]["id"] != current_parent_id:
                        self._move_wwise_object(result[0]["id"], current_parent_id)
                        current_parent_id = result[0]["id"]
                    else:
                        result = self._create_wwise_object(current_parent_id, "WorkUnit", story_soundbank_work_unit_name, {
                            "onNameConflict": "merge"
                        })
                        current_parent_id = result["id"]
                    current_parent_path = f"{current_parent_path}/{story_soundbank_work_unit_name}"

                if is_full_id_voice:
                    story_soundbank_work_unit_name = full_id_work_unit_name
                    result = self._create_wwise_object(current_parent_id, "WorkUnit", story_soundbank_work_unit_name, {
                        "onNameConflict": "merge"
                    })
                    current_parent_id = result["id"]

                soundbank_name = "STO"
                for i in range(1, len(file_name_split) - 1 + container_element_index):
                    file_name_element = file_name_split[i]
                    soundbank_name = f"{soundbank_name}_{file_name_element}"

                if is_full_id_voice:
                    soundbank_name = f"{soundbank_name}_{full_id_work_unit_name}"

                result = self._get_wwise_object(f"where name = \"{soundbank_name}\" and type = \"SoundBank\"")
                if result:
                    soundbank_id = result[0]["id"]
                    soundbank_parent_id = result[0]["parent"]["id"]
                    if soundbank_parent_id != current_parent_id:
                        self._move_wwise_object(soundbank_id, current_parent_id)
                else:
                    result = self._create_wwise_object(current_parent_id, "SoundBank", soundbank_name)
                    soundbank_id = result["id"]
                soundbank_inclusions_array = [{
                    "object": dialog_event_folder_id,
                    "filter": ["events", "structures", "media"]
                }, {
                    "object": sfx_event_folder_id,
                    "filter": ["events", "structures", "media"]
                }]
                self._set_soundbank_inclusions(soundbank_id, soundbank_inclusions_array, "replace")

            dst_sample_path: str = f"{sample_imported_dir_path}/{file_name}"
            self.move_file(original_sample_path, dst_sample_path)
            import_count += 1
            current_count += 1
            self.update_progress_current_count_signal.emit(current_count)

        self.save_wwise_project()

        if is_voice_excel_workbook_modify:
            self.backup_workbook(voice_excel_file_path)
            self.format_voice_workbook(voice_excel_workbook, voice_excel_file_path)
            self.write_workbook(voice_excel_workbook, voice_excel_file_path)

        self.remove_empty_directory(dir_path)

        if len(container_type_invalid_check_list[2]) > 0:
            check_list.append(container_type_invalid_check_list)

        if len(sound_type_invalid_check_list[2]) > 0:
            check_list.append(sound_type_invalid_check_list)

        if len(voice_type_invalid_check_list[2]) > 0:
            check_list.append(voice_type_invalid_check_list)

        if len(check_list) > 0:
            self.show_wwise_object_check_window_signal.emit("素材入库问题检查列表", check_list)

        if self.cancel_job:
            result_str = "素材入库中止"
        else:
            result_str = "素材入库完成"
        if import_count > 0:
            result_str = f"{result_str}\n入库了{import_count}个素材"
        else:
            result_str = f"{result_str}\n没有入库素材"
        if skip_count > 0:
            result_str = f"{result_str}\n跳过了{skip_count}个素材"
        if self.cancel_job or skip_count > 0:
            self.finish_signal.emit("结果", result_str, "warning")
        else:
            self.finish_signal.emit("结果", result_str, "success")

    def _read_voice_excel_data_for_import_job(self, workbook: Workbook) -> dict:
        event_info_dict = {}
        if not workbook:
            return event_info_dict
        for worksheet in workbook.worksheets:
            worksheet_name = worksheet.title
            for i in range(2, worksheet.max_row + 1):
                voice_event = self.get_cell_value(worksheet, i, VoiceExcelUtility.VOICE_EVENT_NAME_COLUMN_INDEX)
                if not voice_event:
                    continue
                dialogue_info = {
                    "worksheet_name": worksheet_name,
                    "row":            i
                }
                if voice_event not in event_info_dict.keys():
                    event_info_dict[voice_event] = []
                event_info_dict[voice_event].append(dialogue_info)
        return event_info_dict

    def _clear_voice_excel_dialogue_record_state_for_import_job(self, workbook: Workbook, event_info_dict: dict, event_name: str) -> bool:
        event_info_list: Optional[list] = event_info_dict.get(event_name, None)
        is_workbook_modify = False
        if not workbook or not event_info_list:
            return False
        for event_info in event_info_list:
            event_info: dict
            worksheet_name: str = event_info["worksheet_name"]
            row: int = event_info["row"]
            worksheet: Worksheet = workbook[worksheet_name]
            current_record_state = self.get_cell_value(worksheet, row, VoiceExcelUtility.RECORD_STATE_COLUMN_INDEX)
            if current_record_state:
                worksheet.cell(row, VoiceExcelUtility.RECORD_STATE_COLUMN_INDEX, "")
                is_workbook_modify = True

        return is_workbook_modify

    def split_work_unit_job(self):
        for character_type in wproj_utility.IMPORT_CHARACTER_TYPE_DICT.values():
            self.update_progress_text_signal.emit(f"处理角色类型...\n{character_type}")
            if self.cancel_job:
                break
            root_actor_mixer_path = f"{ACTOR_MIXER_HIERARCHY}/{character_type}/{character_type}/{character_type}"
            result = self._get_wwise_object(f"\"{root_actor_mixer_path}\"")
            if result:
                root_actor_mixer_id = result[0]["id"]
                result = self._get_wwise_object(f"\"{root_actor_mixer_path}\" select children where type = \"ActorMixer\" and name != \"0 Example\"")
                if result:
                    for character_actor_mixer in result:
                        if self.cancel_job:
                            break
                        character_actor_mixer_id = character_actor_mixer["id"]
                        character_id = character_actor_mixer["name"]
                        character_work_unit_name = f"{character_type} {character_id}"
                        result = self._get_wwise_object(
                                f"from object \"{ACTOR_MIXER_HIERARCHY}/{character_type}\" select descendants where name = \"{character_work_unit_name}\" and type = \"WorkUnit\"")
                        if result:
                            character_work_unit = result[0]
                            if character_work_unit["parent"]["id"] != root_actor_mixer_id:
                                self._move_wwise_object(character_work_unit["id"], root_actor_mixer_id)
                        else:
                            character_work_unit = self._create_wwise_object(root_actor_mixer_id, "WorkUnit", character_work_unit_name)

                        result = self._get_wwise_object(f"\"{root_actor_mixer_path}/{character_work_unit_name}/{character_id}\"")
                        if result:
                            # todo: 工作单元下已有角色混音对象, 无法合并新旧角色混音对象, 需要手工处理
                            pass
                        else:
                            self._move_wwise_object(character_actor_mixer_id, character_work_unit["id"])

            event_root_path = f"{EVENTS}/{character_type}"
            soundbank_folder_root_path = f"{SOUNDBANKS}/{character_type}"

            result = self._get_wwise_object(f"\"{event_root_path}\" select children where type = \"WorkUnit\"")
            if result:
                for character_work_unit in result:
                    if self.cancel_job:
                        break
                    character_id = character_work_unit["name"]
                    character_event_folder_path = f"{event_root_path}/{character_id}/{character_id}"
                    result = self._get_wwise_object(f"\"{character_event_folder_path}\"")
                    if result:
                        character_event_folder_id = result[0]["id"]
                        result = self._get_wwise_object(f"\"{character_event_folder_path}\" select children")
                        if result:
                            for child_folder in result:
                                self._move_wwise_object(child_folder["id"], character_work_unit["id"])
                            result = self._get_wwise_object(f"\"{soundbank_folder_root_path}\" select descendants where name = \"{character_type}{character_id}\" and type = \"SoundBank\"")
                            if result:
                                character_soundbank_id = result[0]["id"]
                                soundbank_inclusions_dict = {
                                    "object": character_event_folder_id,
                                    "filter": ["events", "structures", "media"]
                                }
                                self._set_soundbank_inclusions(character_soundbank_id, [soundbank_inclusions_dict], "remove")
                            self._delete_wwise_object(character_event_folder_id)

            result = self._get_wwise_object(f"\"{soundbank_folder_root_path}\"")
            if result:
                soundbank_folder_id = result[0]["id"]
                result = self._get_wwise_object(f"\"{soundbank_folder_root_path}\" select descendants where type = \"SoundBank\"")
                if result:
                    for character_soundbank in result:
                        if self.cancel_job:
                            break
                        character_soundbank_id = character_soundbank["id"]
                        character_soundbank_name: str = character_soundbank["name"]
                        if character_soundbank_name == character_type or not character_soundbank_name.startswith(character_type):
                            continue
                        character_id = character_soundbank_name[len(character_type):]
                        result = self._get_wwise_object(f"from object \"{soundbank_folder_root_path}\" select descendants where name = \"{character_id}\" and type = \"WorkUnit\"")
                        if result:
                            character_work_unit = result[0]
                            if character_work_unit["parent"]["id"] != soundbank_folder_id:
                                self._move_wwise_object(character_work_unit["id"], soundbank_folder_id)
                        else:
                            character_work_unit = self._create_wwise_object(soundbank_folder_id, "WorkUnit", character_id)

                        if character_soundbank["parent"]["id"] != character_work_unit["id"]:
                            self._move_wwise_object(character_soundbank_id, character_work_unit["id"])

                        result = self._get_wwise_object(f"\"{event_root_path}/{character_id}\"")
                        if result:
                            character_event_work_unit_id = result[0]["id"]
                            result = self._get_soundbank_inclusions(character_soundbank_id)
                            if not result or (result and result[0]["object"] != character_event_work_unit_id):
                                soundbank_inclusions_dict = {
                                    "object": character_event_work_unit_id,
                                    "filter": ["events", "structures", "media"]
                                }
                                self._set_soundbank_inclusions(character_soundbank_id, [soundbank_inclusions_dict], "replace")

        self.save_wwise_project()

        if self.cancel_job:
            result_str = "拆分工作单元中止"
        else:
            result_str = "拆分工作单元完成"

        if self.cancel_job:
            self.finish_signal.emit("结果", result_str, "warning")
        else:
            self.finish_signal.emit("结果", result_str, "success")

    def select_sound_object_within_text_job(self, file_path: str):
        self.update_progress_text_signal.emit(f"读取文本文件...\n{file_path}")
        file = open(file_path)
        text = file.readlines()
        file.close()
        sound_object_id_list: list[str] = []
        sound_object_not_found_list: list[str] = []
        for event_name in text:
            event_name = event_name.strip()
            sound_object_name = event_name.replace("_", " ")
            self.update_progress_text_signal.emit(f"查找声音对象...\n{sound_object_name}")
            result = self._get_wwise_object(f"where name = \"{sound_object_name}\" and type = \"Sound\"")
            if result:
                sound_object_id: str = result[0]["id"]
                if sound_object_id not in sound_object_id_list:
                    sound_object_id_list.append(sound_object_id)
            else:
                sound_object_not_found_list.append(event_name)

        self._show_list_view(sound_object_id_list)

        if len(sound_object_id_list) == 0:
            result_str = "没有查找到任何事件对应的声音对象"
        else:
            result_str = f"查找选中声音对象完成\n查找选中了{len(sound_object_id_list)}个声音对象"
            if len(sound_object_not_found_list) > 0:
                result_str = f"{result_str}\n{len(sound_object_not_found_list)}个事件对应的声音对象没有查找到"

        self.finish_signal.emit("结果", result_str, "success")

    def _get_wwise_object(self, waql: str, return_option_list: Optional[list[str]] = None) -> Optional[list]:
        if not self.is_connected():
            return None
        return_list = ["id", "name", "notes", "parent", "path", "type", "@Color"]
        if return_option_list:
            return_list += return_option_list
        args = {
            "waql":    waql,
            "options": {
                "return": return_list
            }
        }
        result = self.client.call("ak.wwise.core.object.get", args)
        if result:
            return result["return"]
        else:
            return None

    def post_event(self):
        list = self._transport_get_list()
        print(list)
        self.finish_signal.emit("结果", "", "success")

    def _create_wwise_object(self, parent: str, object_type: str, name: str, arg_dict: Optional[dict] = None) -> Optional[dict]:
        if not self.is_connected():
            return None
        args = {
            "parent": parent,
            "type":   object_type,
            "name":   name,
        }
        if arg_dict:
            args |= arg_dict
        result = self.client.call("ak.wwise.core.object.create", args)
        return result

    def _set_wwise_object(self, arg_dict: Optional[dict] = None) -> Optional[dict]:
        if not self.is_connected():
            return None
        result = self.client.call("ak.wwise.core.object.set", arg_dict)
        return result

    def _move_wwise_object(self, wwise_object: str, parent: str):
        if not self.is_connected():
            return
        args = {
            "object": wwise_object,
            "parent": parent
        }
        self.client.call("ak.wwise.core.object.move", args)

    def _copy_wwise_object(self, wwise_object: str, parent: str, on_name_conflict: str = "fail") -> Optional[dict]:
        if not self.is_connected():
            return None
        args = {
            "object":         wwise_object,
            "parent":         parent,
            "onNameConflict": on_name_conflict
        }
        result = self.client.call("ak.wwise.core.object.copy", args)
        return result

    def _rename_wwise_object(self, wwise_object: str, value: str):
        if not self.is_connected():
            return
        args = {
            "object": wwise_object,
            "value":  value
        }
        self.client.call("ak.wwise.core.object.setName", args)

    def _delete_wwise_object(self, wwise_object: str):
        if not self.is_connected():
            return
        args = {
            "object": wwise_object
        }
        self.client.call("ak.wwise.core.object.delete", args)

    def _set_wwise_object_property(self, wwise_object: str, property_name: str, value):
        if not self.is_connected():
            return
        args = {
            "object":   wwise_object,
            "property": property_name,
            "value":    value
        }
        self.client.call("ak.wwise.core.object.setProperty", args)

    def _set_wwise_object_reference(self, wwise_object: str, reference_name: str, value: str):
        if not self.is_connected():
            return
        args = {
            "object":    wwise_object,
            "reference": reference_name,
            "value":     value
        }
        self.client.call("ak.wwise.core.object.setReference", args)

    def _import_audio(self, object_path: str, audio_path: Optional[str] = None, language: str = "SFX", root: Optional[str] = None, import_path: Optional[str] = None,
                      import_operation: str = "useExisting") -> Optional[list[dict]]:
        if not self.is_connected():
            return None
        import_dict = {
            "importLanguage": language,
            "objectPath":     object_path,
            # "objectType":     "Sound"
        }
        if audio_path:
            import_dict["audioFile"] = audio_path
        if root:
            import_dict["importLocation"] = root
        if import_path:
            import_dict["originalsSubFolder"] = import_path
        args = {
            "importOperation": import_operation,
            "imports":         [
                import_dict
            ]
        }
        result = self.client.call("ak.wwise.core.audio.import", args)
        if result:
            return result["objects"]
        else:
            return None

    def _add_switch_container_assignment(self, child: str, state_or_switch: str):
        if not self.is_connected():
            return
        args = {
            "child":         child,
            "stateOrSwitch": state_or_switch
        }
        self.client.call("ak.wwise.core.switchContainer.addAssignment", args)

    def _set_soundbank_inclusions(self, soundbank: str, inclusions: list, operation: Optional[str] = "add"):
        if not self.is_connected():
            return
        args = {
            "soundbank":  soundbank,
            "operation":  operation,
            "inclusions": inclusions
        }
        self.client.call("ak.wwise.core.soundbank.setInclusions", args)

    def _get_soundbank_inclusions(self, soundbank: str, ) -> list:
        if not self.is_connected():
            return []
        args = {
            "soundbank": soundbank
        }
        result = self.client.call("ak.wwise.core.soundbank.getInclusions", args)
        return result["inclusions"]

    def _show_list_view(self, object_list: list[str]):
        if not self.is_connected() or not object_list:
            return
        args = {
            "command": "ShowListView",
            "objects": object_list
        }
        self.client.call("ak.wwise.ui.commands.execute", args)

    def _register_game_obj(self, object_id: int, name: str):
        if not self.is_connected() or object_id is None or not name:
            return
        args = {
            "gameObject": object_id,
            "name":       name
        }
        self.client.call("ak.soundengine.registerGameObj", args)

    def _unregister_game_obj(self, object_id: int):
        if not self.is_connected() or object_id is None:
            return
        args = {
            "gameObject": object_id,
        }
        self.client.call("ak.soundengine.unregisterGameObj", args)

    def _set_default_listeners(self, listeners: list[int]):
        if not self.is_connected() or not listeners:
            return
        args = {
            "listeners": listeners,
        }
        self.client.call("ak.soundengine.setDefaultListeners", args)

    def _post_event(self, event: str, object_id: Optional[int] = None) -> Optional[int]:
        if not self.is_connected() or not event:
            return None
        args: dict = {
            "event": event
        }
        if object_id is not None:
            args.update({
                "gameObject": object_id
            })
        return self.client.call("ak.soundengine.postEvent", args)

    def _stop_playing_id(self, playing_id: int):
        if not self.is_connected() or playing_id is None:
            return
        args = {
            "playingId":          playing_id,
            "transitionDuration": 0,
            "fadeCurve":          4
        }
        self.client.call("ak.soundengine.stopPlayingID", args)

    def _seek_on_event(self, event: str, object_id: int, position: int):
        if not self.is_connected() or not event or object_id is None or position is None:
            return
        args = {
            "event":               event,
            "gameObject":          object_id,
            "position":            position,
            "seekToNearestMarker": False,
            # "playingId":           0
        }
        self.client.call("ak.soundengine.seekOnEvent", args)

    def _ui_get_selected_objects(self) -> list[dict]:
        if not self.is_connected():
            return []
        result = self.client.call("ak.wwise.ui.getSelectedObjects", {})
        return result["objects"]
