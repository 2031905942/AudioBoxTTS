import copy
import math
import re
import traceback
from datetime import datetime

from PySide6.QtCore import QDir, QDirIterator, QFile, QFileInfo, QIODeviceBase, Signal
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.styleable import StyleableObject
from openpyxl.worksheet.worksheet import Worksheet
from pypinyin import Style, lazy_pinyin

from Source.Utility.file_utility import FileUtility


class SyncState:
    """ 台本行同步状态的枚举 """

    NEW: str = "1"
    """ 新增 """

    CHAPTER_CHANGE: str = "2"
    """ 章节/任务有变动 """

    SPEAKER_CHANGE: str = "3"
    """ 人物有变动 """

    ORIGINAL_TEXT_CHANGE: str = "4"
    """ 参照台本有变动 """

    RECORD_TEXT_CHANGE: str = "5"
    """ 录音台本有变动 """


class DialogueRowInfo(object):
    """ 某一行台本的所有数据信息, 用于两张表格做数据比较 """

    def __init__(self):
        self.row: int = 0
        """ 该台本的所在行的序号 """

        self.record_id: str | None = None
        """ 语音ID """

        self.dialogue_id: str | None = None
        """ 游戏内对话的ID """

        self.chapter: str | None = None
        """ 任务/剧情/关卡/章节 """

        self.speaker: str | None = None
        """ 说话人"""

        self.original_text: str | None = None
        """ 参照语言台本 """

        self.record_text: str | None = None
        """ 录音语言台本 """

        self.wwise_event: str | None = None
        """ Wwise语音事件名 """

        self.sync_state: str | None = None
        """ 同步状态信息"""

        self.worksheet_name: str | None = None
        """ 该台本所属工作表的名字 """

        self.cell_info: {
            str: str
            } = {}
        """ 工作簿内的每行的所有单元格数据, key是列的序号, value是该列的单元格内的数据 """

    def clone(self):
        """ 深拷贝 """
        new_dialogue_row_info: DialogueRowInfo = DialogueRowInfo()
        new_dialogue_row_info.row = self.row
        new_dialogue_row_info.worksheet_name = self.worksheet_name
        new_dialogue_row_info.record_id = self.record_id
        new_dialogue_row_info.dialogue_id = self.dialogue_id
        new_dialogue_row_info.chapter = self.chapter
        new_dialogue_row_info.speaker = self.speaker
        new_dialogue_row_info.original_text = self.original_text
        new_dialogue_row_info.record_text = self.record_text
        new_dialogue_row_info.wwise_event = self.wwise_event
        new_dialogue_row_info.sync_state = self.sync_state
        new_dialogue_row_info.cell_info = copy.deepcopy(self.cell_info)
        return new_dialogue_row_info


class VoiceExcelUtility(FileUtility):
    """ 处理录音工作簿的工具类 """

    IGNORE_WORKSHEET_NAME: str = "手工填写"

    VOICE_ID_COLUMN_INDEX: int = 1
    """ 存放语音ID信息的列的序号 """

    DIALOGUE_ID_COLUMN_INDEX: int = 15
    """ 存放台本ID信息(即文案平台上该台本的ID)的列的序号 """

    RECORD_STATE_COLUMN_INDEX: int = 3
    """ 存放录音状态信息的列的序号 """

    SYNC_STATE_COLUMN_INDEX: int = 4
    """ 存放同步状态信息的列的序号 """

    LOCALIZED_LANGUAGE_COLUMN_INDEX: int = 5
    """ 存放本地化语言信息的列的序号 """

    CHAPTER_COLUMN_INDEX: int = 6
    """ 存放章节/任务信息的列的序号 """

    SPEAKER_COLUMN_INDEX: int = 7
    """ 存放人物信息的列的序号 """

    ORIGINAL_TEXT_COLUMN_INDEX: int = 9
    """ 存放参照语言台本信息的列的序号 """

    RECORD_TEXT_COLUMN_INDEX: int = 10
    """ 存放录音语言台本信息的列的序号 """

    VOICE_EVENT_NAME_COLUMN_INDEX: int = 13
    """ 存放语音事件名信息的列的序号 """

    LAST_COLUMN_INDEX: int = DIALOGUE_ID_COLUMN_INDEX
    """ 工作簿最后一列的序号 """

    VOICE_EXCEL_HEADER_NAME_DICT: {
        int: str
        } = {
        1:  "录音ID",
        2:  "录音批次",
        3:  "录音状态\n1: 待录音\n2: 已录音, 待回收\n3. 已回收, 待入库",
        4:  "同步状态\n1: 新增\n2: 章节/任务有变动\n3. 人物有变动\n4. 参照台本有变动\n5. 录音台本有变动",
        5:  "本地化区域",
        6:  "章节/任务",
        7:  "人物",
        8:  "情绪/主题",
        9:  "参照台本",
        10: "录音台本",
        11: "备注",
        12: "问题",
        13: "事件命名",
        14: "文件命名\n(一般用于单一事件多素材. 如无, 重命名读取事件命名)",
        15: "Dialogue ID"
        }
    """ 录音工作簿的工作表的表头名称字典 """

    DIALOGUE_DATA_DIFF_EXCEL_HEADER_NAME_DICT: {
        int: str
        } = {
        1: "行数",
        2: "变动",
        3: "变动前",
        4: "变动后"
        }
    """ 台本数据差异工作簿的工作表的表头名称字典 """

    source_excel_dialogue_data: {
        str: DialogueRowInfo
        } = {}
    """ 源工作簿的台本数据字典 """

    target_excel_dialogue_data: {
        str: DialogueRowInfo
        } = {}
    """ 目标工作簿的台本数据字典 """

    changed_dialogue_data_list: [(str, int, str, str, str)] = []
    """ 同步时有修改的台本行数据列表 """

    source_excel_workbook: Workbook
    """ 源工作簿 """

    target_excel_workbook: Workbook
    """ 目标工作簿 """

    # 信号定义
    show_dialogue_check_list_window_signal = Signal(str, list)

    def __init__(self, voice_job):
        from Source.Job.voice_job import VoiceJob
        voice_job: VoiceJob
        super().__init__(voice_job)

    def sync_dialogue_data(self, source_excel_path: str, target_excel_path: str):
        """
        同步文案平台台本数据到录音工作簿
        :param source_excel_path: 文案平台工作簿的路径
        :param target_excel_path: 录音工作簿的路径
        """
        # self._print_log(f"同步文案平台台本数据. 文案平台工作簿: {source_excel_path}; 录音工作簿: {target_excel_path}.")
        if not self.check_workbook_writability(target_excel_path):
            return
        if not self._read_source_excel_dialogue_data(source_excel_path):
            self.finish_signal.emit("结果", "任务中止", "error")
            return
        if not self._read_target_excel_dialogue_data(target_excel_path):
            self.finish_signal.emit("结果", "任务中止", "error")
            return
        if not self._sync_excel_dialogue_data():
            self.finish_signal.emit("结果", "任务中止", "error")
            return

        if not self._write_voice_excel_dialogue_data(target_excel_path):
            self.finish_signal.emit("结果", "任务中止", "error")
            return

        if not self._create_dialogue_data_diff_workbook(target_excel_path):
            self.finish_signal.emit("结果", "同步文案平台台本数据成功, 创建台本数据差异工作簿失败", "warning")
            return

        self.finish_signal.emit("结果", "同步文案平台台本数据成功, 创建台本数据差异工作簿", "success")

    def sync_localized_dialogue_data(self, source_excel_path: str, target_excel_path: str):
        """
        同步多语言台本数据
        :param source_excel_path: 第一语言录音工作簿的路径
        :param target_excel_path: 第二语言录音工作簿的路径
        """
        # self._print_log(f"同步多语言台本数据. 第一语言录音工作簿: {source_excel_path}; 第二语言录音工作簿: {target_excel_path}.")
        if not self.check_workbook_writability(target_excel_path):
            return
        if not self._read_source_excel_dialogue_data(source_excel_path):
            self.finish_signal.emit("结果", "任务中止", "error")
            return
        if not self._read_target_excel_dialogue_data(target_excel_path):
            self.finish_signal.emit("结果", "任务中止", "error")
            return
        if not self._sync_localized_excel_dialogue_data():
            self.finish_signal.emit("结果", "任务中止", "error")
            return

        if not self._write_voice_excel_dialogue_data(target_excel_path):
            # self._print_log("同步失败.")
            self.finish_signal.emit("结果", "任务中止", "error")
            return

        # self._print_log("同步完成.")
        self.finish_signal.emit("结果", "同步多语言台本数据成功", "success")

    def generate_voice_id_for_voice_excel(self, excel_path: str):
        """
        为录音工作簿生成语音ID
        :param excel_path: 录音工作簿的路径
        """
        # self._print_log(f"为录音工作簿生成语音ID: {excel_path}.")
        if not self.check_workbook_writability(excel_path):
            return
        self.update_progress_text_signal.emit(f"读取录音工作簿...\n\"{excel_path}\"")
        workbook: Workbook = self.load_workbook(excel_path)
        if not workbook:
            # self._print_log_error(f"无法加载工作簿: \"{excel_file_path}\".")
            self.finish_signal.emit("结果", "任务中止", "error")
            return

        if not self.check_worksheet_header_validity(excel_path, workbook):
            self.finish_signal.emit("结果", "任务中止", "error")
            return

        need_saved: bool = False
        self.update_progress_text_signal.emit(f"检查\"说话人\"合法性...")
        illegal_speaker_info_list: [str] = []
        illegal_voice_id_info_list: [str] = []
        dialogue_need_generate_list: [(str, int, str, str)] = []
        # 先完整遍历一遍工作簿以生成语音事件已使用的序号的字典
        speaker_used_voice_id_dict: {
            str: int
            } = {}
        for worksheet in workbook.worksheets:
            worksheet: Worksheet
            for row in range(2, worksheet.max_row + 1):
                row: int
                voice_id: str | None = self.get_cell_value(worksheet, row, VoiceExcelUtility.VOICE_ID_COLUMN_INDEX)
                speaker: str | None = self.get_cell_value(worksheet, row, VoiceExcelUtility.SPEAKER_COLUMN_INDEX)
                if not speaker:
                    for column in range(2, VoiceExcelUtility.LAST_COLUMN_INDEX + 1):
                        column: int
                        if column == VoiceExcelUtility.SPEAKER_COLUMN_INDEX:
                            continue
                        cell_content: str | None = self.get_cell_value(worksheet, row, column)
                        if cell_content:
                            illegal_speaker_info: str = f"\"说话人\"内容为空. 工作表: \"{worksheet.title}\"; 行数: \"{row}\"."
                            illegal_speaker_info_list.append(illegal_speaker_info)
                            break
                    if voice_id:
                        worksheet.cell(row, VoiceExcelUtility.VOICE_ID_COLUMN_INDEX, "")  # 如果说话人为空, 但语音ID存在, 则删掉
                        need_saved = True
                else:
                    speaker_first_letter: [str] = lazy_pinyin(speaker, style=Style.FIRST_LETTER)
                    speaker_abbreviation: str = ""
                    for letter in speaker_first_letter:
                        letter: str
                        letter: str = re.sub("[^a-zA-Z]+", "", letter)
                        if len(letter) == 1:
                            letter = letter.upper()
                        speaker_abbreviation += letter
                    if not speaker_abbreviation:
                        illegal_speaker_info: str = f"\"说话人\"不合法, 请至少包含汉字或英文. 说话人: \"{speaker}\"; 工作表: \"{worksheet.title}\"; 行数: \"{row}\"."
                        illegal_speaker_info_list.append(illegal_speaker_info)
                    else:  # 成功获取说话人内容的缩写
                        if voice_id:  # 如果语音ID内容也存在, 则检查其格式是否符合
                            illegal: bool = False
                            if len(voice_id) > len(speaker_abbreviation):
                                id_string: str = voice_id[len(speaker_abbreviation):]
                                if not voice_id.startswith(speaker_abbreviation) or not id_string.isnumeric():
                                    illegal = True
                                else:
                                    id_number: int = int(id_string)
                                    if speaker_abbreviation not in speaker_used_voice_id_dict.keys():
                                        speaker_used_voice_id_dict[speaker_abbreviation] = id_number
                                    elif id_number > speaker_used_voice_id_dict[speaker_abbreviation]:
                                        speaker_used_voice_id_dict[speaker_abbreviation] = id_number
                            else:
                                illegal = True
                            if illegal:
                                illegal_speaker_info: str = f"\"语音ID\"与\"说话人\"不匹配, 自动删除该行的\"语音ID\". 录音ID: \"{voice_id}\"; 说话人: \"{speaker}\"; 工作表: \"{worksheet.title}\"; 行数: \"{row}\"."
                                worksheet.cell(row, VoiceExcelUtility.VOICE_ID_COLUMN_INDEX, "")
                                dialogue_need_generate_list.append((worksheet.title, row, speaker_abbreviation, speaker))
                                need_saved = True
                                illegal_voice_id_info_list.append(illegal_speaker_info)  # print(f"工作表: {worksheet.title}; 行数: {row}; 说话人\"{speaker}\": {speaker_abbreviation}.")
                        else:
                            dialogue_need_generate_list.append((worksheet.title, row, speaker_abbreviation, speaker))

        if len(illegal_speaker_info_list) > 0:
            illegal_info_list: [(str, bool)] = []
            for illegal_speaker_info in illegal_speaker_info_list:
                illegal_info_list.append((illegal_speaker_info, False))
            for illegal_voice_id_info in illegal_voice_id_info_list:
                illegal_info_list.append((illegal_voice_id_info, True))

            self.show_dialogue_check_list_window_signal.emit("工作簿问题", illegal_info_list)

            if need_saved:
                self.format_voice_workbook(workbook, excel_path)
                if not self.write_workbook(workbook, excel_path):
                    self.finish_signal.emit("结果", "任务中止, 请先排查完工作簿的问题再执行", "error")
                else:
                    self.finish_signal.emit("结果", "任务中止, 请先排查完工作簿的问题再执行", "warning")
            else:
                self.finish_signal.emit("结果", "任务中止, 请先排查完工作簿的问题再执行", "warning")
        else:
            if len(illegal_voice_id_info_list) > 0:
                illegal_info_list: [(str, bool)] = []
                for illegal_voice_id_info in illegal_voice_id_info_list:
                    illegal_info_list.append((illegal_voice_id_info, True))
                self.show_dialogue_check_list_window_signal.emit("工作簿问题", illegal_info_list)
            if len(dialogue_need_generate_list) > 0:
                if not self.backup_workbook(excel_path):
                    self.finish_signal.emit("结果", "任务中止", "error")
                    return False

                self.update_progress_text_signal.emit(f"生成语音ID...")
                for dialogue_need_generate in dialogue_need_generate_list:
                    dialogue_need_generate: (str, int, str, str)
                    worksheet_name: str = dialogue_need_generate[0]
                    row: int = dialogue_need_generate[1]
                    speaker_abbreviation: str = dialogue_need_generate[2]
                    new_id: int = 0
                    if speaker_abbreviation in speaker_used_voice_id_dict.keys():
                        new_id = speaker_used_voice_id_dict[speaker_abbreviation] + 1
                    speaker_used_voice_id_dict[speaker_abbreviation] = new_id
                    new_voice_id: str = f"{speaker_abbreviation}{new_id}"
                    worksheet: Worksheet = workbook[worksheet_name]
                    worksheet.cell(row, VoiceExcelUtility.VOICE_ID_COLUMN_INDEX, new_voice_id)  # print(f"生成新语音ID\"{new_voice_id}\". 工作表: {worksheet_name}; 行数: {row}.")

                self.format_voice_workbook(workbook, excel_path)
                if not self.write_workbook(workbook, excel_path):
                    self.finish_signal.emit("结果", "任务中止", "error")

                if not self._create_generate_voice_id_result_workbook(excel_path, dialogue_need_generate_list):
                    self.finish_signal.emit("结果", "生成语音ID成功, 创建生成结果工作簿失败", "warning")
                    return False

                self.finish_signal.emit("结果", "生成语音ID成功, 创建生成结果工作簿", "success")

    def generate_voice_event_for_voice_excel(self, excel_path: str):
        """
        为录音工作簿生成语音事件
        :param excel_path: 录音工作簿的路径
        """
        # print(f"[voice_record_excel_util] 加载录音工作簿: \"{excel_path}\".")
        self.update_progress_text_signal.emit(f"读取录音工作簿...\n\"{excel_path}\"")
        workbook: Workbook = self.load_workbook(excel_path)
        if not workbook:
            self.finish_signal.emit("结果", "任务中止", "error")
            return

        if not self.check_worksheet_header_validity(excel_path, workbook):
            self.finish_signal.emit("结果", "任务中止", "error")
            return

        chapter_abbreviation_used_no_dict: {
            str: int
            } = {}
        dialogue_row_need_to_generate_voice_event_list: [(str, int, str)] = []
        dialogue_check_list: [str] = []
        worksheet: Worksheet
        # 先完整遍历一遍工作簿以生成语音事件已使用的序号的字典
        self.update_progress_text_signal.emit(f"检查\"章节/任务\"合法性...")
        for worksheet in workbook.worksheets:
            worksheet: Worksheet
            worksheet_title: str = worksheet.title
            if VoiceExcelUtility.IGNORE_WORKSHEET_NAME in worksheet_title:
                continue

            for row in range(2, worksheet.max_row + 1):
                row: int
                voice_event_name: str | None = self.get_cell_value(worksheet, row, VoiceExcelUtility.VOICE_EVENT_NAME_COLUMN_INDEX)
                if voice_event_name:
                    voice_event_name_element: [str] = voice_event_name.split("_")
                    if not voice_event_name_element[-1].isnumeric():
                        dialogue_check_str: str = f"事件名\"{voice_event_name}\"不合法, 最后的元素不是纯数字, 请更正事件. 工作表: \"{worksheet.title}\"; 行数: \"{row}\"."
                        dialogue_check_list.append(dialogue_check_str)
                        continue
                    voice_event_name_no: int = int(voice_event_name_element[-1])
                    voice_event_name_element.pop()
                    voice_event_name_element.pop()
                    chapter_abbreviation: str = "_".join(voice_event_name_element)
                    # print(f"[voice_record_excel_util] 事件名: \"{voice_event_name}\"; 序号: \"{voice_event_name_no}\".")
                    if chapter_abbreviation in chapter_abbreviation_used_no_dict:
                        if voice_event_name_no > chapter_abbreviation_used_no_dict[chapter_abbreviation]:
                            chapter_abbreviation_used_no_dict[chapter_abbreviation] = voice_event_name_no
                    else:
                        chapter_abbreviation_used_no_dict[chapter_abbreviation] = voice_event_name_no
                else:
                    voice_event_prefix: str = ""
                    version: str = ""
                    chapter: str = ""
                    quest_abbreviation: str = ""

                    if "-" in worksheet_title:
                        version_split_element_list: [str] = list(filter(None, worksheet_title.split("-")))
                        if len(version_split_element_list) > 0 and version_split_element_list[0].startswith("A"):
                            version = version_split_element_list[0]
                        if len(version_split_element_list) > 1:
                            chapter = version_split_element_list[1]

                    if chapter:
                        chapter_split_element_list: [str] = list(filter(None, chapter.split("_")))
                        if len(chapter_split_element_list) > 0:
                            chapter = chapter_split_element_list[0]
                            chapter = chapter.capitalize()

                    quest: str | None = self.get_cell_value(worksheet, row, VoiceExcelUtility.CHAPTER_COLUMN_INDEX)
                    if quest:
                        quest_first_letter: [str] = lazy_pinyin(quest, style=Style.FIRST_LETTER)

                        for letter in quest_first_letter:
                            letter: str
                            letter: str = re.sub("[^a-zA-Z0-9]+", "", letter)
                            if len(letter) == 1:
                                letter = letter.upper()
                            quest_abbreviation += letter
                    else:
                        dialogue_check_str: str = f"\"章节/任务\"内容为空, 无法生成语音事件. 工作表: \"{worksheet.title}\"; 行数: \"{row}\"."
                        dialogue_check_list.append(dialogue_check_str)

                    if quest_abbreviation:
                        voice_event_prefix = f"VO"
                    if version:
                        voice_event_prefix = f"{voice_event_prefix}_{version}"
                    if chapter:
                        voice_event_prefix = f"{voice_event_prefix}_{chapter}"
                    if quest_abbreviation:
                        voice_event_prefix = f"{voice_event_prefix}_{quest_abbreviation}"
                    if voice_event_prefix:
                        dialogue_row_need_to_generate_voice_event: (str, int, str) = (worksheet_title, row, voice_event_prefix)
                        dialogue_row_need_to_generate_voice_event_list.append(dialogue_row_need_to_generate_voice_event)

        # 如果有台本存在问题(如数据缺失等), 任务中止
        if len(dialogue_check_list) > 0:
            dialogue_check_info_list: [(str, bool)] = []
            for dialogue_check in dialogue_check_list:
                dialogue_check: str
                dialogue_check_info_list.append((dialogue_check, False))
            self.show_dialogue_check_list_window_signal.emit("工作簿问题", dialogue_check_info_list)
            self.error_signal.emit(f"请先排查完工作簿问题后再执行任务.")
            self.finish_signal.emit("结果", "任务中止", "error")
            return

        if not self.backup_workbook(excel_path):
            self.finish_signal.emit("结果", "任务中止", "error")
            return

        self.update_progress_text_signal.emit(f"生成语音事件...")
        for dialogue_row_need_to_generate_voice_event in dialogue_row_need_to_generate_voice_event_list:
            dialogue_row_need_to_generate_voice_event: (str, int, str)
            worksheet_title: str = dialogue_row_need_to_generate_voice_event[0]
            row: int = dialogue_row_need_to_generate_voice_event[1]
            voice_event_prefix: str = dialogue_row_need_to_generate_voice_event[2]
            if voice_event_prefix in chapter_abbreviation_used_no_dict.keys():
                new_no: int = chapter_abbreviation_used_no_dict[voice_event_prefix] + 1
            else:
                new_no: int = 0
            chapter_abbreviation_used_no_dict[voice_event_prefix] = new_no
            # 每50句语音分在一个声音库中
            bank_no: int = math.floor(new_no / 50)
            voice_event: str = f"{voice_event_prefix}_{bank_no}_{new_no}"
            # self._print_log(f"工作表: \"{worksheet_title}\"; 行数: \"{row}\"; 语音事件: \"{voice_event}\".")
            worksheet: Worksheet = workbook[worksheet_title]
            worksheet.cell(row, VoiceExcelUtility.VOICE_EVENT_NAME_COLUMN_INDEX, voice_event)

        self.format_voice_workbook(workbook, excel_path)

        if not self.write_workbook(workbook, excel_path):
            self.finish_signal.emit("结果", "任务中止", "error")
            return

        self.finish_signal.emit("结果", "生成语音事件成功", "success")

    def format_voice_excel(self, excel_path: str) -> bool:
        """
        格式化录音工作簿
        :param excel_path: 录音工作簿的路径
        :returns: 是否成功
        """
        # self._print_log(f"格式化录音工作簿: {excel_path}.")
        if not self.check_workbook_writability(excel_path):
            return False
        workbook: Workbook = self.load_workbook(excel_path)
        if not workbook:
            # self._print_log_error(f"无法加载工作簿: \"{excel_file_path}\".")
            self.finish_signal.emit("结果", "任务中止", "error")
            return False

        # if not self._check_worksheet_header_validity(excel_path, workbook):
        #     self.finish_signal.emit("结果", "任务中止", "error")
        #     return False

        self.format_voice_workbook(workbook, excel_path)
        if not self.write_workbook(workbook, excel_path):
            self.finish_signal.emit("结果", "任务中止", "error")
            return False

        self.finish_signal.emit("结果", "格式化工作簿成功", "success")
        return True

    def rename_voice_sample_by_voice_excel(self, dir_path: str, excel_path: str):
        qdir: QDir = QDir(dir_path)
        if not qdir.exists():
            self.error_signal.emit(f"语音素材目录路径不合法:\n{dir_path}")
            self.finish_signal.emit("结果", "任务中止", "error")
            return

        self.update_progress_text_signal.emit(f"读取录音工作簿...\n\"{excel_path}\"")
        workbook: Workbook = self.load_workbook(excel_path)
        if not workbook:
            self.finish_signal.emit("结果", "任务中止", "error")
            return

        if not self.check_worksheet_header_validity(excel_path, workbook):
            self.finish_signal.emit("结果", "任务中止", "error")
            return

        voice_sample_path_info_dict: {
            str: [str]
            } = {}
        voice_sample_check_list: [(str, bool)] = []
        invalid_voice_sample_name: [str] = []

        # 先遍历语音素材
        it: QDirIterator = QDirIterator(qdir, QDirIterator.IteratorFlag.Subdirectories)
        while it.hasNext():
            it.next()
            if not it.fileInfo().isFile() or not it.fileName().lower().endswith(".wav"):
                continue
            voice_sample_name: str = it.fileInfo().baseName()
            voice_sample_path: str = it.filePath()
            if voice_sample_name not in voice_sample_path_info_dict:
                voice_sample_path_info_dict[voice_sample_name] = [voice_sample_path]
            else:
                if voice_sample_path not in voice_sample_path_info_dict[voice_sample_name]:
                    voice_sample_path_info_dict[voice_sample_name].append(voice_sample_path)
                    voice_sample_check_str: (str, bool) = (
                        f"目录下存在多个同名的语音素材\"{it.fileName()}\", 请检查是否相同的文件. 路径: \"" + "; ".join(voice_sample_path_info_dict[voice_sample_name]) + "\".", False)
                    voice_sample_check_list.append(voice_sample_check_str)
                    if voice_sample_name not in invalid_voice_sample_name:
                        invalid_voice_sample_name.append(voice_sample_name)

        for voice_sample_name in invalid_voice_sample_name:
            voice_sample_name: str
            voice_sample_path_info_dict.pop(voice_sample_name)

        if len(voice_sample_check_list) > 0:
            self.show_dialogue_check_list_window_signal.emit("语音素材问题", voice_sample_check_list)
            self.error_signal.emit(f"请先排查完语音素材问题后再执行任务.")
            self.finish_signal.emit("结果", "任务中止", "error")
            return

        if len(voice_sample_path_info_dict) == 0:
            self.finish_signal.emit("结果", "没有语音素材需要重命名", "warning")
            return

        need_to_save_workbook: bool = False
        rename_voice_sample_dir_path: str = f"{dir_path}/重命名语音素材"
        original_voice_sample_dir_path: str = f"{dir_path}/重命名成功的原语音素材"

        if not self.backup_directory(dir_path):
            self.finish_signal.emit("结果", "任务中止", "error")
            return

        self.update_progress_text_signal.emit(f"重命名语音素材...")
        rename_voice_sample_count: int = 0
        for worksheet in workbook.worksheets:
            worksheet: Worksheet
            for row in range(2, worksheet.max_row + 1):
                row: int
                voice_id: str | None = self.get_cell_value(worksheet, row, VoiceExcelUtility.VOICE_ID_COLUMN_INDEX)
                if not voice_id:
                    continue

                if voice_id not in voice_sample_path_info_dict.keys():
                    continue

                original_voice_sample_path: str = voice_sample_path_info_dict[voice_id][0]

                voice_event: str | None = self.get_cell_value(worksheet, row, VoiceExcelUtility.VOICE_EVENT_NAME_COLUMN_INDEX)

                if not voice_event:
                    continue

                rename_voice_sample_name: str = voice_event.replace("_", " ")
                rename_voice_sample_name = f"{rename_voice_sample_name}.wav"
                localized_language: str | None = self.get_cell_value(worksheet, row, VoiceExcelUtility.LOCALIZED_LANGUAGE_COLUMN_INDEX)
                if localized_language:
                    rename_voice_sample_dir_path = f"{rename_voice_sample_dir_path}-{localized_language}"

                rename_voice_sample_path = f"{rename_voice_sample_dir_path}/{rename_voice_sample_name}"

                if not self.copy_file(original_voice_sample_path, rename_voice_sample_path):
                    self.finish_signal.emit("结果", "任务中止", "error")
                    return

                worksheet.cell(row, VoiceExcelUtility.RECORD_STATE_COLUMN_INDEX, "3")
                need_to_save_workbook = True

                rename_voice_sample_count += 1

                voice_sample_path_info_dict.pop(voice_id)

                if not self.move_file(original_voice_sample_path, f"{original_voice_sample_dir_path}/{voice_id}.wav"):
                    self.finish_signal.emit("结果", "任务中止", "error")
                    return

        voice_sample_not_found_dir_path = f"{dir_path}/未找到的语音素材"
        for voice_sample_name, voice_sample_path_list in voice_sample_path_info_dict.items():
            voice_sample_name: str
            voice_sample_path_list: [str]
            voice_sample_path: str = voice_sample_path_list[0]
            if not self.move_file(voice_sample_path, f"{voice_sample_not_found_dir_path}/{voice_sample_name}.wav"):
                self.finish_signal.emit("结果", "任务中止", "error")
                return

        self.remove_empty_directory(dir_path)

        if need_to_save_workbook:
            self.backup_workbook(excel_path)
            self.format_voice_workbook(workbook, excel_path)
            if not self.write_workbook(workbook, excel_path):
                self.finish_signal.emit("结果", "任务中止", "error")
                return
            self.finish_signal.emit("结果", f"重命名语音素材成功\n重命名成功{rename_voice_sample_count}个, 未找到{len(voice_sample_path_info_dict)}个", "success")
        else:
            self.finish_signal.emit("结果", f"没有在录音工作簿上找到相应的语音素材\n未找到{len(voice_sample_path_info_dict)}个", "warning")

    def _read_source_excel_dialogue_data(self, excel_path: str) -> bool:
        """
        读取源工作簿的台本数据
        :param excel_path: 工作簿的路径
        :returns: 是否成功
        """
        self.source_excel_dialogue_data.clear()
        self.update_progress_text_signal.emit(f"读取源工作簿...\n\"{excel_path}\"")
        # self._print_log(f"读取源工作簿: \"{excel_path}\".")
        self.source_excel_workbook = self.load_workbook(excel_path)
        if not self.source_excel_workbook:
            # self._print_log_error(f"无法加载源工作簿: \"{excel_path}\".")
            return False
        if not self.source_excel_workbook.worksheets or len(self.source_excel_workbook.worksheets) == 0:
            self.error_signal.emit(f"工作簿内没有工作表: \"{excel_path}\".")
            # self._print_log_error(f"源工作簿内没有工作表: \"{excel_path}\".")
            return False

        dialogue_check_list: [str] = []

        for worksheet in self.source_excel_workbook.worksheets:
            # self._print_log(f"读取工作表: \"{worksheet.title}\".")
            row: int = 2
            for i in range(2, worksheet.max_row + 1):
                dialogue_id = self.get_cell_value(worksheet, i, VoiceExcelUtility.DIALOGUE_ID_COLUMN_INDEX)
                if not dialogue_id:
                    continue

                dialogue_row_info = DialogueRowInfo()
                dialogue_row_info.row = row
                dialogue_row_info.dialogue_id = dialogue_id
                dialogue_row_info.worksheet_name = worksheet.title

                voice_id = self.get_cell_value(worksheet, i, VoiceExcelUtility.VOICE_ID_COLUMN_INDEX)
                if voice_id:
                    dialogue_row_info.record_id = voice_id

                chapter = self.get_cell_value(worksheet, i, VoiceExcelUtility.CHAPTER_COLUMN_INDEX)
                if chapter:
                    dialogue_row_info.chapter = chapter

                speaker = self.get_cell_value(worksheet, i, VoiceExcelUtility.SPEAKER_COLUMN_INDEX)
                if speaker:
                    dialogue_row_info.speaker = speaker
                else:
                    dialogue_check_str: str = f"\"说话人\"内容为空. 工作表: \"{worksheet.title}\"; 行数: \"{row}\"."
                    dialogue_check_list.append(dialogue_check_str)  # self._print_log(f"{dialogue_check_str}")

                original_text = self.get_cell_value(worksheet, i, VoiceExcelUtility.ORIGINAL_TEXT_COLUMN_INDEX)
                if original_text:
                    dialogue_row_info.original_text = original_text
                else:
                    dialogue_check_str: str = f"\"参照台本\"内容为空. 工作表: \"{worksheet.title}\"; 行数: \"{row}\"."
                    dialogue_check_list.append(dialogue_check_str)  # self._print_log(f"{dialogue_check_str}")

                record_text = self.get_cell_value(worksheet, i, VoiceExcelUtility.RECORD_TEXT_COLUMN_INDEX)
                if record_text:
                    dialogue_row_info.record_text = record_text
                else:
                    dialogue_check_str: str = f"\"录音台本\"内容为空. 工作表: \"{worksheet.title}\"; 行数: \"{row}\"."
                    dialogue_check_list.append(dialogue_check_str)  # self._print_log(f"{dialogue_check_str}")

                wwise_event: str = self.get_cell_value(worksheet, i, VoiceExcelUtility.VOICE_EVENT_NAME_COLUMN_INDEX)
                if wwise_event:
                    dialogue_row_info.wwise_event = wwise_event
                if not (chapter or speaker or original_text or record_text):
                    continue

                row += 1

                self.source_excel_dialogue_data[dialogue_id] = dialogue_row_info
        # self._print_log(f"读取源工作簿完成.")
        if len(dialogue_check_list) > 0:
            dialogue_check_info_list: [(str, bool)] = []
            for dialogue_check in dialogue_check_list:
                dialogue_check: str
                dialogue_check_info_list.append((dialogue_check, False))
            self.show_dialogue_check_list_window_signal.emit("工作簿问题", dialogue_check_info_list)
        return True

    def _read_target_excel_dialogue_data(self, excel_path: str) -> bool:
        """
        读取目标工作簿的台本数据
        :param excel_path: 工作簿的路径
        :returns: 是否成功
        """
        self.target_excel_dialogue_data.clear()
        self.update_progress_text_signal.emit(f"读取目标工作簿...\n\"{excel_path}\"")
        # self._print_log(f"读取目标工作簿: \"{excel_path}\".")
        self.target_excel_workbook = self.load_workbook(excel_path)
        if not self.target_excel_workbook:
            # self._print_log_error(f"无法加载目标工作簿: \"{excel_path}\".")
            return False
        if not self.target_excel_workbook.worksheets or len(self.target_excel_workbook.worksheets) == 0:
            self.error_signal.emit(f"工作簿内没有工作表: \"{excel_path}\".")
            # self._print_log_error(f"目标工作簿内没有工作表: \"{excel_path}\".")
            return False

        if not self.check_worksheet_header_validity(excel_path, self.target_excel_workbook):
            return False

        for worksheet in self.target_excel_workbook.worksheets:
            if VoiceExcelUtility.IGNORE_WORKSHEET_NAME in worksheet.title:
                continue

            # self._print_log(f"读取工作表: \"{worksheet.title}\".")
            for i in range(2, worksheet.max_row + 1):
                dialogue_row_info = DialogueRowInfo()
                dialogue_row_info.row = i
                dialogue_row_info.worksheet_name = worksheet.title

                voice_id: str | None = self.get_cell_value(worksheet, i, VoiceExcelUtility.VOICE_ID_COLUMN_INDEX)
                if voice_id:
                    dialogue_row_info.record_id = voice_id

                dialogue_id: str | None = self.get_cell_value(worksheet, i, VoiceExcelUtility.DIALOGUE_ID_COLUMN_INDEX)
                if dialogue_id:
                    dialogue_row_info.dialogue_id = dialogue_id

                sync_state: str | None = self.get_cell_value(worksheet, i, VoiceExcelUtility.SYNC_STATE_COLUMN_INDEX)
                if sync_state:
                    dialogue_row_info.sync_state = sync_state

                chapter: str | None = self.get_cell_value(worksheet, i, VoiceExcelUtility.CHAPTER_COLUMN_INDEX)
                if chapter:
                    dialogue_row_info.chapter = chapter

                speaker: str | None = self.get_cell_value(worksheet, i, VoiceExcelUtility.SPEAKER_COLUMN_INDEX)
                if speaker:
                    dialogue_row_info.speaker = speaker

                original_text: str | None = self.get_cell_value(worksheet, i, VoiceExcelUtility.ORIGINAL_TEXT_COLUMN_INDEX)
                if original_text:
                    dialogue_row_info.original_text = original_text

                record_text: str | None = self.get_cell_value(worksheet, i, VoiceExcelUtility.RECORD_TEXT_COLUMN_INDEX)
                if record_text:
                    dialogue_row_info.record_text = record_text

                wwise_event: str | None = self.get_cell_value(worksheet, i, VoiceExcelUtility.VOICE_EVENT_NAME_COLUMN_INDEX)
                if wwise_event:
                    dialogue_row_info.wwise_event = wwise_event

                if not (dialogue_id or chapter or speaker or original_text or record_text):
                    continue

                # 针对录音台本表, 需要将每一行的所有单元格数据都存储一下, 因为后面相当于对整张录音台本表重写一遍
                for j in range(1, VoiceExcelUtility.LAST_COLUMN_INDEX + 1):
                    cell_value = self.get_cell_value(worksheet, i, j)
                    if not cell_value:
                        continue
                    dialogue_row_info.cell_info[str(j)] = cell_value

                self.target_excel_dialogue_data[dialogue_id] = dialogue_row_info
        # self._print_log(f"读取目标工作簿完成.")
        return True

    def _sync_excel_dialogue_data(self) -> bool:
        """
        同步文案平台工作簿和录音工作簿中的台本数据
        :return: 是否成功
        """
        self.update_progress_text_signal.emit(f"同步台本数据...")
        # self._print_log("同步台本数据.")
        VoiceExcelUtility.changed_dialogue_data_list.clear()
        if len(self.source_excel_dialogue_data) == 0:
            return False
        for dialogue_id in self.source_excel_dialogue_data:
            source_data: DialogueRowInfo = self.source_excel_dialogue_data[dialogue_id]
            if dialogue_id in self.target_excel_dialogue_data:  # 如果该对话ID已经在录音台本表里了
                target_data: DialogueRowInfo = self.target_excel_dialogue_data[dialogue_id]

                target_data.row = source_data.row
                target_data.worksheet_name = source_data.worksheet_name
                # 如果需要同步的数据项有修改, 则需要标注下"2" 提醒同步后需要关注一下该行, 有内容改动, 并且将该台本数据加入到有变化的数据的列表中记录
                sync_state_list: [str] = []
                if target_data.sync_state:
                    sync_state_list = target_data.sync_state.split("|")
                if source_data.chapter and target_data.chapter != source_data.chapter:
                    if SyncState.CHAPTER_CHANGE not in sync_state_list:
                        sync_state_list.append(SyncState.CHAPTER_CHANGE)
                    changed_dialogue_data: (str, int, str, str, str) = (
                        source_data.worksheet_name, source_data.row, VoiceExcelUtility.VOICE_EXCEL_HEADER_NAME_DICT[6], target_data.chapter, source_data.chapter, source_data.dialogue_id)
                    VoiceExcelUtility.changed_dialogue_data_list.append(changed_dialogue_data)

                target_speaker: str | None = target_data.speaker
                save_speaker: str = ""
                if target_speaker and "#" in target_speaker:
                    target_speaker = target_speaker[0:target_speaker.index("#")]
                    save_speaker = target_data.speaker[target_data.speaker.index("#") + 1:]
                if source_data.speaker and target_speaker != source_data.speaker:
                    if SyncState.SPEAKER_CHANGE not in sync_state_list:
                        sync_state_list.append(SyncState.SPEAKER_CHANGE)
                    changed_dialogue_data: (str, int, str, str, str) = (
                        source_data.worksheet_name, source_data.row, VoiceExcelUtility.VOICE_EXCEL_HEADER_NAME_DICT[7], target_speaker, source_data.speaker, source_data.dialogue_id)
                    VoiceExcelUtility.changed_dialogue_data_list.append(changed_dialogue_data)

                if target_data.original_text != source_data.original_text:
                    if SyncState.ORIGINAL_TEXT_CHANGE not in sync_state_list:
                        sync_state_list.append(SyncState.ORIGINAL_TEXT_CHANGE)
                    changed_dialogue_data: (str, int, str, str, str) = (
                        source_data.worksheet_name, source_data.row, VoiceExcelUtility.VOICE_EXCEL_HEADER_NAME_DICT[9], target_data.original_text, source_data.original_text, source_data.dialogue_id)
                    VoiceExcelUtility.changed_dialogue_data_list.append(changed_dialogue_data)

                if target_data.record_text != source_data.record_text:
                    if SyncState.RECORD_TEXT_CHANGE not in sync_state_list:
                        sync_state_list.append(SyncState.RECORD_TEXT_CHANGE)
                    changed_dialogue_data: (str, int, str, str, str) = (
                        source_data.worksheet_name, source_data.row, VoiceExcelUtility.VOICE_EXCEL_HEADER_NAME_DICT[10], target_data.record_text, source_data.record_text, source_data.dialogue_id)
                    VoiceExcelUtility.changed_dialogue_data_list.append(changed_dialogue_data)

                sync_state_list.sort()

                target_data.sync_state = "|".join(sync_state_list)

                if source_data.chapter and source_data.chapter != "":
                    target_data.chapter = source_data.chapter
                if save_speaker:
                    target_data.speaker = source_data.speaker + "#" + save_speaker
                else:
                    if source_data.speaker and source_data.speaker != "":
                        target_data.speaker = source_data.speaker
                target_data.original_text = source_data.original_text
                target_data.record_text = source_data.record_text

            else:
                dialogue_row_info = source_data.clone()
                dialogue_row_info.sync_state = SyncState.NEW  # 新增的台本行标"1"

                self.target_excel_dialogue_data[dialogue_id] = dialogue_row_info

        # 遍历一次录音台本表的数据, 如果对话ID不存在在对话台本表里, 则代表可能该对话已经被删除了
        dialogue_id_to_delete = [x for x in self.target_excel_dialogue_data.keys() if x not in self.source_excel_dialogue_data]
        for dialogue_id in dialogue_id_to_delete:
            deleted_dialogue_data: DialogueRowInfo = self.target_excel_dialogue_data[dialogue_id]
            changed_dialogue_data: (str, int, str, str, str) = (
                deleted_dialogue_data.worksheet_name, deleted_dialogue_data.row, "删除", deleted_dialogue_data.dialogue_id, deleted_dialogue_data.original_text)
            VoiceExcelUtility.changed_dialogue_data_list.append(changed_dialogue_data)
            del self.target_excel_dialogue_data[dialogue_id]

        for old_dialogue_row_info in self.target_excel_dialogue_data.values():
            wwise_event: str = old_dialogue_row_info.wwise_event
            if wwise_event:
                duplicate_dialogue_id_list: [str] = self._find_dialogue_id_by_wwise_event(self.source_excel_dialogue_data, wwise_event)
                if len(duplicate_dialogue_id_list) > 1:
                    self._print_log(f"源台本表存在重复Wwise事件名的台本行. Wwise事件名: \"{old_dialogue_row_info.wwise_event}\".")

        # 最后对录音台本数据按照行做排序
        self.target_excel_dialogue_data = dict(sorted(self.target_excel_dialogue_data.items(), key=lambda item: item[1].row))

        return True

    def _sync_localized_excel_dialogue_data(self) -> bool:
        """
        同步多语言录音工作簿中的台本数据
        :return: 是否成功
        """
        if len(self.source_excel_dialogue_data) == 0:
            return False
        self.update_progress_text_signal.emit(f"同步台本数据...")
        # self._print_log("同步台本数据.")
        for dialogue_id in self.source_excel_dialogue_data:
            source_data: DialogueRowInfo = self.source_excel_dialogue_data[dialogue_id]
            if dialogue_id in self.target_excel_dialogue_data:  # 如果该对话ID已经在录音台本表里了
                target_data: DialogueRowInfo = self.target_excel_dialogue_data[dialogue_id]
                target_data.record_id = source_data.record_id
                target_data.wwise_event = source_data.wwise_event

        # 最后对台本数据按照行做排序
        self.target_excel_dialogue_data = dict(sorted(self.target_excel_dialogue_data.items(), key=lambda item: item[1].row))
        return True

    def _find_dialogue_id_by_wwise_event(self, dialogue_data: {str, DialogueRowInfo}, wwise_event: str) -> [str]:
        new_dialogue_id_list: [str] = []
        if not dialogue_data:
            return new_dialogue_id_list
        for item in dialogue_data.items():
            dialogue_row_info: DialogueRowInfo = item[1]
            if wwise_event == dialogue_row_info.wwise_event:
                if item[0] not in new_dialogue_id_list:
                    new_dialogue_id_list.append(item[0])
        return new_dialogue_id_list

    def _write_voice_excel_dialogue_data(self, write_excel_path: str) -> bool:
        """
        台本数据写入录音工作簿
        :param write_excel_path: 工作簿的路径
        :return: 是否成功
        """
        # self._print_log(f"写入录音工作簿: \"{write_excel_path}\".")
        if not self.target_excel_workbook:
            return False

        if not self.backup_workbook(write_excel_path):
            return False

        self.update_progress_text_signal.emit(f"写入录音工作簿...\n\"{write_excel_path}\"")
        existed_work_sheet_name_list: [str] = []
        for dialogue_id in self.target_excel_dialogue_data:
            dialogue_row_info = self.target_excel_dialogue_data[dialogue_id]
            worksheet_name = dialogue_row_info.worksheet_name
            if worksheet_name not in existed_work_sheet_name_list:
                existed_work_sheet_name_list.append(worksheet_name)
            if worksheet_name not in self.target_excel_workbook.sheetnames:
                self._print_log(f"录音工作簿内创建新工作表: {worksheet_name}")
                for existed_worksheet in self.target_excel_workbook.worksheets:
                    worksheet = self.target_excel_workbook.copy_worksheet(existed_worksheet)
                    worksheet.title = worksheet_name
                    break
        delete_work_sheet_name_list: [str] = list(set(self.target_excel_workbook.sheetnames).difference(existed_work_sheet_name_list))

        for sheet_name in delete_work_sheet_name_list:
            if VoiceExcelUtility.IGNORE_WORKSHEET_NAME not in sheet_name:
                self.target_excel_workbook.remove_sheet(self.target_excel_workbook[sheet_name])

        for existed_worksheet in self.target_excel_workbook.worksheets:
            if VoiceExcelUtility.IGNORE_WORKSHEET_NAME in existed_worksheet.title:
                continue
            existed_worksheet.delete_rows(2, existed_worksheet.max_row)
            if existed_worksheet.max_column > VoiceExcelUtility.LAST_COLUMN_INDEX:
                existed_worksheet.delete_cols(VoiceExcelUtility.LAST_COLUMN_INDEX + 1, existed_worksheet.max_column)
            row: int = 2
            for dialogue_id in self.target_excel_dialogue_data:
                dialogue_row_info = self.target_excel_dialogue_data[dialogue_id]
                if dialogue_row_info.worksheet_name != existed_worksheet.title:
                    continue

                for column in dialogue_row_info.cell_info:
                    existed_worksheet.cell(row, int(column), dialogue_row_info.cell_info[column])
                existed_worksheet.cell(row, VoiceExcelUtility.DIALOGUE_ID_COLUMN_INDEX, dialogue_id)
                existed_worksheet.cell(row, VoiceExcelUtility.VOICE_ID_COLUMN_INDEX, dialogue_row_info.record_id)
                existed_worksheet.cell(row, VoiceExcelUtility.CHAPTER_COLUMN_INDEX, dialogue_row_info.chapter)
                if dialogue_row_info.sync_state != 0:
                    existed_worksheet.cell(row, VoiceExcelUtility.SYNC_STATE_COLUMN_INDEX, dialogue_row_info.sync_state)
                else:
                    existed_worksheet.cell(row, VoiceExcelUtility.SYNC_STATE_COLUMN_INDEX, "")
                existed_worksheet.cell(row, VoiceExcelUtility.SPEAKER_COLUMN_INDEX, dialogue_row_info.speaker)
                existed_worksheet.cell(row, VoiceExcelUtility.ORIGINAL_TEXT_COLUMN_INDEX, dialogue_row_info.original_text)
                existed_worksheet.cell(row, VoiceExcelUtility.RECORD_TEXT_COLUMN_INDEX, dialogue_row_info.record_text)
                existed_worksheet.cell(row, VoiceExcelUtility.VOICE_EVENT_NAME_COLUMN_INDEX, dialogue_row_info.wwise_event)
                row += 1

        self.format_voice_workbook(self.target_excel_workbook, write_excel_path)

        if self.write_workbook(self.target_excel_workbook, write_excel_path):
            # self._print_log(f"写入录音工作簿完成.")
            return True
        else:
            return False

    def _create_dialogue_data_diff_workbook(self, workbook_path: str) -> bool:
        """
        创建台本数据差异工作簿
        :param workbook_path: 录音工作簿的路径
        :return: 是否成功
        """
        if len(VoiceExcelUtility.changed_dialogue_data_list) == 0:
            return True
        voice_workbook_file_info: QFileInfo = QFileInfo(workbook_path)
        voice_workbook_parent_path: str = voice_workbook_file_info.path()
        current_time = datetime.now()
        current_time_string: str = current_time.strftime("%m-%d %H-%M-%S")
        dialogue_data_diff_workbook_path: str = f"{voice_workbook_parent_path}/台本数据差异 {current_time_string}.xlsx"
        self.update_progress_text_signal.emit(f"创建台本数据差异工作簿...\n\"{dialogue_data_diff_workbook_path}\"")
        dialogue_data_diff_workbook: Workbook = Workbook()
        worksheet_name_list: [str] = []

        for changed_dialogue_data in VoiceExcelUtility.changed_dialogue_data_list:
            changed_dialogue_data: (str, int, str, str, str)
            worksheet_name: str = changed_dialogue_data[0]
            if worksheet_name not in worksheet_name_list:
                worksheet_name_list.append(worksheet_name)

        default_worksheet: Worksheet = dialogue_data_diff_workbook.active
        # 创建工作表
        for worksheet_name in worksheet_name_list:
            worksheet_name: str
            dialogue_data_diff_workbook.create_sheet(worksheet_name)
        dialogue_data_diff_workbook.remove(default_worksheet)
        # noinspection PyProtectedMember,PyUnresolvedReferences
        dialogue_data_diff_workbook._sheets.sort(key=lambda sheet: sheet.title)

        thin: Side = Side(border_style="thin", color="000000")
        for worksheet in dialogue_data_diff_workbook.worksheets:
            worksheet: Worksheet

            # 创建表头
            for column in range(1, 5):
                column: int
                cell: StyleableObject | Cell = worksheet.cell(1, column)
                cell.value = VoiceExcelUtility.DIALOGUE_DATA_DIFF_EXCEL_HEADER_NAME_DICT[column]
                cell.font = Font(name="Microsoft YaHei", size=13, color="000000", bold=True)
                cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")
                cell.fill = PatternFill("solid", fgColor="ffff00")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

            row = 2
            for changed_dialogue_data in VoiceExcelUtility.changed_dialogue_data_list:
                changed_dialogue_data: (str, int, str, str, str)
                worksheet_name: str = changed_dialogue_data[0]
                if worksheet_name != worksheet.title:
                    continue

                for column in range(1, 5):
                    column: int
                    cell: StyleableObject | Cell = worksheet.cell(row, column)
                    cell.value = changed_dialogue_data[column]
                    cell.font = Font(name="Microsoft YaHei", size=13, color="000000")
                    cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")
                    cell.fill = PatternFill("solid", fgColor="fde9d9")
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

                row += 1

            worksheet.column_dimensions["A"].width = 20
            worksheet.column_dimensions["B"].width = 30
            worksheet.column_dimensions["C"].width = 100
            worksheet.column_dimensions["D"].width = 100

        VoiceExcelUtility.changed_dialogue_data_list = []
        if not self.write_workbook(dialogue_data_diff_workbook, dialogue_data_diff_workbook_path):
            self.error_signal.emit(f"创建台本数据差异工作簿失败:\n\"{dialogue_data_diff_workbook_path}\"")
            return False

        return True

    def _create_generate_voice_id_result_workbook(self, workbook_path: str, generate_info_list: [(str, int, str, str)]) -> bool:
        """
        创建语音ID生成结果工作簿
        :param workbook_path: 录音工作簿的路径
        :param generate_info_list: 语音ID生成信息列表
        :return: 是否成功
        """
        # print(f"voice_workbook_parent_path: \"{voice_workbook_parent_path}\".")
        voice_workbook_file_info: QFileInfo = QFileInfo(workbook_path)
        voice_workbook_parent_path: str = voice_workbook_file_info.path()
        current_time = datetime.now()
        current_time_string: str = current_time.strftime("%m-%d %H-%M-%S")
        generate_voice_id_result_workbook_path: str = f"{voice_workbook_parent_path}/生成语音ID结果 {current_time_string}.xlsx"
        self.update_progress_text_signal.emit(f"创建语音ID生成结果工作簿...\n\"{generate_voice_id_result_workbook_path}\"")
        generate_voice_id_result_workbook: Workbook = Workbook()
        worksheet: Worksheet = generate_voice_id_result_workbook.active
        worksheet.title = "默认"
        generate_voice_id_result_dict: {
            str: {
                str: int
                }
            } = {}
        speaker_count_dict: {
            str: int
            } = {}
        worksheet_count_dict: {
            str: int
            } = {}
        for generate_info in generate_info_list:
            generate_info: (str, int, str, str)
            worksheet_name: str = generate_info[0]
            speaker: str = generate_info[3]
            if worksheet_name not in worksheet_count_dict:
                worksheet_count_dict[worksheet_name] = 1
            else:
                worksheet_count_dict[worksheet_name] += 1

            if speaker not in speaker_count_dict:
                speaker_count_dict[speaker] = 1
            else:
                speaker_count_dict[speaker] += 1

            if worksheet_name not in generate_voice_id_result_dict.keys():
                generate_voice_id_result_dict[worksheet_name] = {}
                generate_voice_id_result_dict[worksheet_name][speaker] = 1
            else:
                if speaker not in generate_voice_id_result_dict[worksheet_name]:
                    generate_voice_id_result_dict[worksheet_name][speaker] = 1
                else:
                    generate_voice_id_result_dict[worksheet_name][speaker] += 1

        column: int = 2
        for worksheet_name, speaker_info in generate_voice_id_result_dict.items():
            worksheet.cell(1, column, worksheet_name)
            for speaker, count in speaker_info.items():
                row = 2
                while self.get_cell_value(worksheet, row, 1):
                    current_speaker: str = self.get_cell_value(worksheet, row, 1)
                    if speaker == current_speaker:
                        break
                    row += 1
                worksheet.cell(row, 1, speaker)
                worksheet.cell(row, column, str(count))

            column += 1

        total_count_row: int = len(speaker_count_dict.keys()) + 2
        total_count_column: int = len(worksheet_count_dict.keys()) + 2
        for row in range(2, total_count_row + 1):
            speaker: str | None = self.get_cell_value(worksheet, row, 1)
            if speaker and speaker in speaker_count_dict:
                worksheet.cell(row, total_count_column, str(speaker_count_dict[speaker]))

        for column in range(2, total_count_column + 1):
            column: int
            worksheet_name: str | None = self.get_cell_value(worksheet, 1, column)
            if worksheet_name and worksheet_name in worksheet_count_dict:
                worksheet.cell(total_count_row, column, str(worksheet_count_dict[worksheet_name]))

        worksheet.cell(1, 1, "生成数量")
        worksheet.cell(total_count_row, 1, "合计")
        worksheet.cell(1, total_count_column, "合计")
        worksheet.cell(total_count_row, total_count_column, str(len(generate_info_list)))

        thin: Side = Side(border_style="thin", color="000000")
        # 表头格式化
        for column in range(1, worksheet.max_column + 1):
            column: int
            cell: StyleableObject | Cell = worksheet.cell(1, column)
            cell.font = Font(name="Microsoft YaHei", size=13, color="000000", bold=True)
            cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="ffff00")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for row in range(2, worksheet.max_row + 1):
            cell: StyleableObject | Cell = worksheet.cell(row, 1)
            cell.font = Font(name="Microsoft YaHei", size=13, color="000000", bold=False)
            cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="c5d9f1")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for column in range(2, worksheet.max_column + 1):
                cell: StyleableObject | Cell = worksheet.cell(row, column)
                cell.font = Font(name="Microsoft YaHei", size=13, color="000000", bold=False)
                cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")
                cell.fill = PatternFill("solid", fgColor="fde9d9")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max(len(str(cell.value)) * 2, 20)
        for col, value in dims.items():
            worksheet.column_dimensions[col].width = value

        if not self.write_workbook(generate_voice_id_result_workbook, generate_voice_id_result_workbook_path):
            self.error_signal.emit(f"创建语音ID生成结果工作簿失败:\n\"{generate_voice_id_result_workbook_path}\"")
            return False
        return True

    def format_voice_workbook(self, workbook: Workbook, workbook_file_path: str):
        """
        格式化录音工作簿
        :param workbook: 录音工作簿
        :param workbook_file_path: 录音工作簿的路径
        """
        self.update_progress_text_signal.emit(f"格式化录音工作簿...\n\"{workbook_file_path}\"")
        thin: Side = Side(border_style="thin", color="000000")
        sync_state_pattern_fill: PatternFill = PatternFill("solid", fgColor="ff3300")
        default_pattern_fill: PatternFill = PatternFill("solid", fgColor="ffffff")
        voice_state_pattern_fill1: PatternFill = PatternFill("solid", fgColor="f4b084")
        voice_state_pattern_fill2: PatternFill = PatternFill("solid", fgColor="ffff00")
        voice_state_pattern_fill3: PatternFill = PatternFill("solid", fgColor="92d050")

        for worksheet in workbook.worksheets:
            worksheet: Worksheet
            for row_index in range(1, worksheet.max_row + 1):
                row = worksheet.row_dimensions[row_index]
                row.fill = PatternFill()
                row.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for column_key in worksheet.column_dimensions.keys():
                column_key: str
                column = worksheet.column_dimensions[column_key]
                column.fill = PatternFill()
                column.border = Border(left=thin, right=thin, top=thin, bottom=thin)

            if worksheet.max_column > VoiceExcelUtility.LAST_COLUMN_INDEX:
                worksheet.delete_cols(VoiceExcelUtility.LAST_COLUMN_INDEX + 1, worksheet.max_column)

            worksheet.freeze_panes = "A2"

            # 表头格式化
            for i in range(1, VoiceExcelUtility.LAST_COLUMN_INDEX + 1):
                i: int
                cell: StyleableObject | Cell = worksheet.cell(1, i)
                cell.value = VoiceExcelUtility.VOICE_EXCEL_HEADER_NAME_DICT[i]
                cell.font = Font(name="Microsoft YaHei", size=13, color="000000", bold=True)
                cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")
                cell.fill = PatternFill("solid", fgColor="ffff00")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for worksheet in workbook.worksheets:
            for row in reversed(range(2, worksheet.max_row + 1)):
                row: int
                original_text: str | None = self.get_cell_value(worksheet, row, VoiceExcelUtility.ORIGINAL_TEXT_COLUMN_INDEX)
                if not original_text:
                    worksheet.delete_rows(row)

            for row in range(2, worksheet.max_row + 1):
                row: int
                record_state: str | None = self.get_cell_value(worksheet, row, VoiceExcelUtility.RECORD_STATE_COLUMN_INDEX)
                sync_state: str | None = self.get_cell_value(worksheet, row, VoiceExcelUtility.SYNC_STATE_COLUMN_INDEX)

                voice_state_pattern_fill: PatternFill = default_pattern_fill

                sync_state_list: [str] = []
                if sync_state:
                    sync_state_list = sync_state.split("|")

                if record_state == "1":
                    voice_state_pattern_fill = voice_state_pattern_fill1
                elif record_state == "2":
                    voice_state_pattern_fill = voice_state_pattern_fill2
                elif record_state == "3":
                    voice_state_pattern_fill = voice_state_pattern_fill3
                for column in range(1, VoiceExcelUtility.LAST_COLUMN_INDEX + 1):
                    column: int
                    cell: StyleableObject | Cell = worksheet.cell(row, column)
                    cell.font = Font(name="Microsoft YaHei", size=13, color="000000", bold=False)
                    cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")
                    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    if len(sync_state_list) > 0 and (
                            (column == VoiceExcelUtility.CHAPTER_COLUMN_INDEX) and (SyncState.CHAPTER_CHANGE in sync_state_list) or (column == VoiceExcelUtility.SPEAKER_COLUMN_INDEX) and (
                            SyncState.SPEAKER_CHANGE in sync_state_list) or (column == VoiceExcelUtility.ORIGINAL_TEXT_COLUMN_INDEX and SyncState.ORIGINAL_TEXT_CHANGE in sync_state_list) or (
                                    column == VoiceExcelUtility.RECORD_TEXT_COLUMN_INDEX) and (SyncState.RECORD_TEXT_CHANGE in sync_state_list)):
                        cell.fill = sync_state_pattern_fill
                    elif 1 <= column <= VoiceExcelUtility.RECORD_TEXT_COLUMN_INDEX and record_state:
                        cell.fill = voice_state_pattern_fill
                    elif 1 <= column <= VoiceExcelUtility.CHAPTER_COLUMN_INDEX - 1:
                        cell.fill = PatternFill("solid", fgColor="c5d9f1")
                    elif VoiceExcelUtility.CHAPTER_COLUMN_INDEX <= column <= VoiceExcelUtility.RECORD_TEXT_COLUMN_INDEX:
                        cell.fill = PatternFill("solid", fgColor="fde9d9")
                    elif column == VoiceExcelUtility.RECORD_TEXT_COLUMN_INDEX + 1:
                        cell.fill = PatternFill("solid", fgColor="e6b8b7")
                    elif column == VoiceExcelUtility.RECORD_TEXT_COLUMN_INDEX + 2:
                        cell.fill = PatternFill("solid", fgColor="ccc0da")
                    elif VoiceExcelUtility.VOICE_EVENT_NAME_COLUMN_INDEX <= column <= VoiceExcelUtility.DIALOGUE_ID_COLUMN_INDEX:
                        cell.fill = PatternFill("solid", fgColor="d8e4bc")

            worksheet.auto_filter.ref = worksheet.dimensions
        # noinspection PyProtectedMember,PyUnresolvedReferences
        workbook._sheets.sort(key=lambda sheet: sheet.title)
        workbook.active = 0

    def load_workbook(self, file_path: str) -> Workbook | None:
        """
        读取工作簿
        :param file_path: 工作簿的路径
        :return: 工作簿
        """
        if not file_path:
            self.error_signal.emit("读取工作簿失败, 路径为空.")
            # self._print_log_error(f"工作簿路径为空.")
            return None
        self.update_progress_text_signal.emit(f"读取工作簿...\n\"{file_path}\"")
        try:
            work_book = load_workbook(file_path)
        except Exception as error:
            self.error_signal.emit(f"读取工作簿发生异常: {error}.")
            self._print_log_error(f"读取工作簿发生异常: {traceback.format_exc()}.")
            return None
        return work_book

    def write_workbook(self, work_book: Workbook, file_path: str) -> bool:
        """
        写入工作簿
        :param work_book: 工作簿
        :param file_path: 工作簿的路径
        :return: 是否成功
        """
        self.update_progress_text_signal.emit(f"写入工作簿...\n\"{file_path}\"")
        if not work_book:
            self.error_signal.emit(f"写入工作簿失败, 工作簿为空.")
            return False
        if not file_path:
            self.error_signal.emit(f"写入工作簿失败, 路径为空.")
            return False
        try:
            work_book.save(file_path)
        except Exception as error:
            self.error_signal.emit(f"写入工作簿发生异常: {error}.")
            self._print_log_error(f"写入工作簿发生异常: {traceback.format_exc()}.")
            return False
        return True

    def backup_workbook(self, file_path: str) -> bool:
        """
        备份工作簿
        :param file_path: 原工作簿的路径
        :return: 是否成功
        """
        self.update_progress_text_signal.emit(f"备份工作簿...\n\"{file_path}\"")
        backup_excel_path = file_path
        backup_excel_path = backup_excel_path[:-5]

        current_time = datetime.now()
        current_time_string: str = current_time.strftime("%m-%d %H-%M-%S")
        backup_excel_path = f"{backup_excel_path}(备份) {current_time_string}.xlsx"
        backup_workbook = self.load_workbook(file_path)
        # self._print_log(f"原工作簿备份: \"{backup_excel_path}\".")
        if not self.write_workbook(backup_workbook, backup_excel_path):
            self.error_signal.emit(f"工作簿备份失败: \"{backup_excel_path}\".")
            return False
        return True

    def check_worksheet_header_validity(self, workbook_path: str, workbook: Workbook) -> bool:
        """
        检查录音工作簿的工作表的表头的合法性
        :param workbook_path: 录音工作簿的路径
        :param workbook: 录音工作簿
        :return: 是否合法
        """
        validity: bool = True
        self.update_progress_text_signal.emit(f"检查语音工作簿的表头的合法性...\n\"{workbook_path}\"")
        for worksheet in workbook.worksheets:
            worksheet: Worksheet

            for column_index, column_header_name in VoiceExcelUtility.VOICE_EXCEL_HEADER_NAME_DICT.items():
                column_index: int
                column_header_name: str
                header_cell_value: str | None = self.get_cell_value(worksheet, 1, column_index)
                if not header_cell_value or header_cell_value != column_header_name:
                    self.error_signal.emit(
                            f"工作簿表头不合法: \"{workbook_path}\".\n工作表: \"{worksheet.title}\"; 列数: \"{column_index}\";\n正确表头: \"{column_header_name}\"; 当前表头: \"{header_cell_value}\".")
                    validity = False

        return validity

    def get_cell_value(self, worksheet: Worksheet, row: int, column: int) -> str | None:
        """
        获取单元格的值
        :param worksheet: 工作表
        :param row: 行
        :param column: 列
        :return: 单元格的值
        """
        if not worksheet:
            return None
        value = worksheet.cell(row, column).value
        if value and not isinstance(value, str):
            value: str = str(value).strip()
        return value

    def check_workbook_writability(self, workbook_path: str) -> bool:
        """
        检查工作簿的可写性
        :param workbook_path: 工作簿的路径
        :return: 是否可写
        """
        if not workbook_path or not QFile.exists(workbook_path):
            self.error_signal.emit(f"无法读取工作簿. 路径不合法: \"{workbook_path}\".")
            return False

        self.update_progress_text_signal.emit(f"检查工作簿的可写性...\n\"{workbook_path}\"")
        file: QFile = QFile(workbook_path)
        try:
            if not file.open(QIODeviceBase.OpenModeFlag.ReadWrite):
                self.finish_signal.emit("警告", f"任务中止, 工作簿不可写入, 请先解除对工作簿的占用再执行.\n工作簿路径\"{workbook_path}\".", "warning")
                file.close()
                return False
            else:
                file.close()
                return True
        except Exception as error:
            self.finish_signal.emit("错误", f"任务中止, 读取工作簿发生异常: \"{error}\".", "error")
            return False
